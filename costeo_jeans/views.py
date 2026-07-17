from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.views.generic import DetailView
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.contrib import messages
from django.http import JsonResponse
import openpyxl
from openpyxl.utils import get_column_letter
from core.auth_utils import es_diseno, es_administracion
from productos.models import ReferenciaColor, Producto
from .models import Insumo, MovimientoInsumo, Proceso, Confeccionista, Costeo, CostoFijo, DetalleCostoFijo, DetalleCantidadTalla
from .forms import (
    InsumoForm, MovimientoInsumoForm, ProcesoForm, ConfeccionistaForm, CostoFijoForm,
    CosteoModelForm, DetalleInsumoFormSet, DetalleProcesoFormSet, AjustePrecioForm, DetalleTallaFormSet,
    DetalleInsumo, DetalleProceso, TarifaConfeccionista, TarifaConfeccionistaForm,
)


def _es_diseno_o_admin(user):
    return es_diseno(user) or es_administracion(user)


_rol_requerido = user_passes_test(_es_diseno_o_admin, login_url='core:acceso_denegado')


class CosteoRolMixin(UserPassesTestMixin):
    """Mismo patrón de clientes/views.py: acceso restringido a Diseño/Administración."""
    def test_func(self):
        return _es_diseno_o_admin(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para acceder a esta sección de Costeo.")
        return redirect('core:index')


class PanelCosteoView(LoginRequiredMixin, CosteoRolMixin, TemplateView):
    template_name = 'costeo_jeans/panel_costeo.html'

class InsumoListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = Insumo
    template_name = 'costeo_jeans/insumo_list.html'
    context_object_name = 'insumos'
    paginate_by = 10

    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Insumo.objects.filter(empresa=empresa_actual).order_by('nombre')
        return Insumo.objects.none()




class InsumoCreateView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = Insumo
    form_class = InsumoForm
    template_name = 'costeo_jeans/insumo_form.html'
    success_url = reverse_lazy('costeo_jeans:insumo_list')
    
    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'tenant', None)
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Añadir Nuevo Insumo'
        context['boton'] = 'Guardar Insumo'
        return context
    
    
    
    
    
    
class InsumoUpdateView(LoginRequiredMixin, CosteoRolMixin, UpdateView):
    model = Insumo
    form_class = InsumoForm
    template_name = 'costeo_jeans/insumo_form.html'
    success_url = reverse_lazy('costeo_jeans:insumo_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Insumo.objects.filter(empresa=empresa_actual)
        return Insumo.objects.none()
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Insumo'
        context['boton'] = 'Actualizar Insumo'
        return context
class InsumoDeleteView(LoginRequiredMixin, CosteoRolMixin, DeleteView):
    model = Insumo
    template_name = 'costeo_jeans/insumo_confirm_delete.html'
    success_url = reverse_lazy('costeo_jeans:insumo_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Insumo.objects.filter(empresa=empresa_actual)
        return Insumo.objects.none()
class ProcesoListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = Proceso
    template_name = 'costeo_jeans/proceso_list.html'
    context_object_name = 'procesos'
    paginate_by = 10
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Proceso.objects.filter(empresa=empresa_actual).order_by('nombre')
        return Proceso.objects.none()
class ProcesoCreateView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = Proceso
    form_class = ProcesoForm
    template_name = 'costeo_jeans/proceso_form.html'
    success_url = reverse_lazy('costeo_jeans:proceso_list')
    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'tenant', None)
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Añadir Nuevo Proceso'
        context['boton'] = 'Guardar Proceso'
        return context
class ProcesoUpdateView(LoginRequiredMixin, CosteoRolMixin, UpdateView):
    model = Proceso
    form_class = ProcesoForm
    template_name = 'costeo_jeans/proceso_form.html'
    success_url = reverse_lazy('costeo_jeans:proceso_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Proceso.objects.filter(empresa=empresa_actual)
        return Proceso.objects.none()
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Proceso'
        context['boton'] = 'Actualizar Proceso'
        return context
class ProcesoDeleteView(LoginRequiredMixin, CosteoRolMixin, DeleteView):
    model = Proceso
    template_name = 'costeo_jeans/proceso_confirm_delete.html'
    success_url = reverse_lazy('costeo_jeans:proceso_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Proceso.objects.filter(empresa=empresa_actual)
        return Proceso.objects.none()
class ConfeccionistaListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = Confeccionista
    template_name = 'costeo_jeans/confeccionista_list.html'
    context_object_name = 'confeccionistas'
    paginate_by = 10
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Confeccionista.objects.filter(empresa=empresa_actual).order_by('nombre')
        return Confeccionista.objects.none()
class ConfeccionistaCreateView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = Confeccionista
    form_class = ConfeccionistaForm
    template_name = 'costeo_jeans/confeccionista_form.html'
    success_url = reverse_lazy('costeo_jeans:confeccionista_list')
    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'tenant', None)
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Añadir Nuevo Confeccionista'
        context['boton'] = 'Guardar Confeccionista'
        return context
class ConfeccionistaUpdateView(LoginRequiredMixin, CosteoRolMixin, UpdateView):
    model = Confeccionista
    form_class = ConfeccionistaForm
    template_name = 'costeo_jeans/confeccionista_form.html'
    success_url = reverse_lazy('costeo_jeans:confeccionista_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Confeccionista.objects.filter(empresa=empresa_actual)
        return Confeccionista.objects.none()
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Confeccionista'
        context['boton'] = 'Actualizar Confeccionista'
        return context
class ConfeccionistaDeleteView(LoginRequiredMixin, CosteoRolMixin, DeleteView):
    model = Confeccionista
    template_name = 'costeo_jeans/confeccionista_confirm_delete.html'
    success_url = reverse_lazy('costeo_jeans:confeccionista_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return Confeccionista.objects.filter(empresa=empresa_actual)
        return Confeccionista.objects.none()
class CostoFijoListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = CostoFijo
    template_name = 'costeo_jeans/costofijo_list.html'
    context_object_name = 'costos_fijos'
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return CostoFijo.objects.filter(empresa=empresa_actual).order_by('nombre')
        return CostoFijo.objects.none()
class CostoFijoCreateView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = CostoFijo
    form_class = CostoFijoForm
    template_name = 'costeo_jeans/costofijo_form.html'
    success_url = reverse_lazy('costeo_jeans:costofijo_list')
    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'tenant', None)
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Añadir Nuevo Costo Fijo'
        context['boton'] = 'Guardar Costo'
        return context
class CostoFijoUpdateView(LoginRequiredMixin, CosteoRolMixin, UpdateView):
    model = CostoFijo
    form_class = CostoFijoForm
    template_name = 'costeo_jeans/costofijo_form.html'
    success_url = reverse_lazy('costeo_jeans:costofijo_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return CostoFijo.objects.filter(empresa=empresa_actual)
        return CostoFijo.objects.none()
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Costo Fijo'
        context['boton'] = 'Actualizar Costo'
        return context
class CostoFijoDeleteView(LoginRequiredMixin, CosteoRolMixin, DeleteView):
    model = CostoFijo
    template_name = 'costeo_jeans/costofijo_confirm_delete.html'
    success_url = reverse_lazy('costeo_jeans:costofijo_list')
    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return CostoFijo.objects.filter(empresa=empresa_actual)
        return CostoFijo.objects.none()
    

@login_required
@_rol_requerido
def costeo_create_step1(request):
    empresa_actual = getattr(request, 'tenant', None)
    if request.method == 'POST':
        form = CosteoModelForm(request.POST, empresa=empresa_actual)
        # Formset validado contra una instancia sin guardar todavía -- solo
        # para chequear la forma de los datos antes de tocar la base de datos.
        talla_formset = DetalleTallaFormSet(request.POST, instance=Costeo(empresa=empresa_actual), prefix='tallas')
        if form.is_valid() and talla_formset.is_valid() and empresa_actual:
            with transaction.atomic():
                costeo = form.save(commit=False)
                costeo.empresa = empresa_actual
                costeo.save()
                talla_formset.instance = costeo
                talla_formset.save()
                costeo.cantidad_producida = costeo.detalle_tallas.aggregate(total=Sum('cantidad'))['total'] or 0
                costeo.save(update_fields=['cantidad_producida'])
            return redirect('costeo_jeans:costeo_create_step2', costeo_id=costeo.id)
        elif not empresa_actual:
            form.add_error(None, "Error: No se pudo determinar la empresa para tu usuario.")
    else:
        form = CosteoModelForm(empresa=empresa_actual)
        talla_formset = DetalleTallaFormSet(instance=Costeo(empresa=empresa_actual), prefix='tallas')
    return render(request, 'costeo_jeans/costeo_create_step1.html', {'form': form, 'talla_formset': talla_formset})


@login_required
@_rol_requerido
def costeo_create_step2(request, costeo_id):
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(Costeo, id=costeo_id, empresa=empresa_actual)
    
    insumos = Insumo.objects.filter(empresa=empresa_actual)
    insumo_categorias = {insumo.id: insumo.categoria for insumo in insumos}   
    
    if request.method == 'POST':
        insumo_formset = DetalleInsumoFormSet(request.POST, instance=costeo, prefix='insumos', form_kwargs={'empresa': empresa_actual})
        proceso_formset = DetalleProcesoFormSet(request.POST, instance=costeo, prefix='procesos', form_kwargs={'empresa': empresa_actual})
        if insumo_formset.is_valid() and proceso_formset.is_valid():
            with transaction.atomic():
                insumos = insumo_formset.save()
                procesos = proceso_formset.save()
                
                
                
                subtotal_variable = 0
                for insumo in insumos:
                    subtotal_variable += insumo.costo_total
                for proceso in procesos:
                    subtotal_variable += proceso.costo_total
                costos_fijos_defecto = CostoFijo.objects.filter(empresa=empresa_actual, incluir_por_defecto=True)
                for cf in costos_fijos_defecto:
                    valor_calculado = 0
                    if cf.tipo == CostoFijo.TipoCosto.PORCENTAJE_SUBTOTAL:
                        valor_calculado = subtotal_variable * (cf.valor / 100)
                    elif cf.tipo == CostoFijo.TipoCosto.VALOR_FIJO_UNIDAD:
                        valor_calculado = cf.valor * costeo.cantidad_producida
                    DetalleCostoFijo.objects.create(
                        costeo=costeo,
                        costo_fijo=cf,
                        valor_calculado=valor_calculado
                    )
            return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)
    else:
        insumo_formset = DetalleInsumoFormSet(instance=costeo, prefix='insumos', form_kwargs={'empresa': empresa_actual})
        proceso_formset = DetalleProcesoFormSet(instance=costeo, prefix='procesos', form_kwargs={'empresa': empresa_actual})
        
        # --- INICIO DE LA MODIFICACIÓN ---
    # 1. Obtener todos los insumos de la empresa
    insumos = Insumo.objects.filter(empresa=empresa_actual)
    # 2. Crear un diccionario: {id_insumo: "CATEGORIA"}
    insumo_categorias = {insumo.id: insumo.categoria for insumo in insumos}
    # --- FIN DE LA MODIFICACIÓN ---    
    
    context = {
        'costeo': costeo,
        'insumo_formset': insumo_formset,
        'proceso_formset': proceso_formset,
        'insumo_categorias_json': json.dumps(insumo_categorias)
    }
    return render(request, 'costeo_jeans/costeo_create_step2.html', context)


@login_required
@_rol_requerido
def costeo_summary(request, costeo_id):
    empresa_actual = getattr(request, 'tenant', None)

    costeo = get_object_or_404(
        Costeo.objects.select_related('referencia_color').prefetch_related(
            'detalle_insumos__insumo',
            'detalle_tallas',
            # Precargamos la tarifa, y a través de ella, el proceso y el confeccionista
            'detalle_procesos__tarifa__proceso',
            'detalle_procesos__tarifa__confeccionista',
            'detalle_costos_fijos__costo_fijo'
        ),
        id=costeo_id,
        empresa=empresa_actual
    )
    esta_finalizado = costeo.estado == Costeo.EstadoCosteo.FINALIZADO

    if request.method == 'POST':
        if esta_finalizado:
            messages.error(request, "Este costeo está Finalizado y no se puede editar. Reábrelo primero si necesitas ajustarlo.")
            return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)
        ajuste_form = AjustePrecioForm(request.POST, instance=costeo)
        if ajuste_form.is_valid():
            ajuste_form.save()
            return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)
    else:
        ajuste_form = AjustePrecioForm(instance=costeo)

    context = {
        'costeo': costeo,
        'ajuste_form': ajuste_form,
        'esta_finalizado': esta_finalizado,
        'puede_reabrir': esta_finalizado and es_administracion(request.user),
    }
    return render(request, 'costeo_jeans/costeo_summary.html', context)


@login_required
@_rol_requerido
def api_buscar_referencia_color(request):
    """Autocompletar (Select2) de referencia+color reales del catálogo, para enlazar un costeo."""
    empresa_actual = getattr(request, 'tenant', None)
    term = request.GET.get('term', '').strip()
    if not empresa_actual or len(term) < 1:
        return JsonResponse({'results': []})

    referencias = ReferenciaColor.objects.filter(
        Q(referencia_base__icontains=term) | Q(color__icontains=term),
        empresa=empresa_actual
    ).order_by('referencia_base', 'color')[:20]

    return JsonResponse({'results': [
        {'id': rc.pk, 'text': f"{rc.referencia_base} - {rc.color or 'Sin color'}"} for rc in referencias
    ]})


@login_required
@_rol_requerido
def costeo_actualizar_catalogo(request, costeo_id):
    """
    Actualiza el costo/precio de venta de todas las tallas (Producto) de la
    referencia+color enlazada, con el resultado de este costeo. Nunca
    automático: siempre muestra antes/después y exige confirmación explícita
    (el mismo criterio usado esta sesión para no sobreescribir precios
    reales en lote sin que alguien lo vea y lo apruebe primero). Solo se
    permite sobre un costeo Finalizado -- no tiene sentido empujar al
    catálogo un costo que todavía es un borrador en progreso.
    """
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(
        Costeo.objects.select_related('referencia_color'), id=costeo_id, empresa=empresa_actual
    )

    if costeo.estado != Costeo.EstadoCosteo.FINALIZADO:
        messages.error(request, "Solo puedes actualizar el catálogo con un costeo Finalizado.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    if not costeo.referencia_color:
        messages.error(request, "Este costeo no está enlazado a ninguna referencia del catálogo -- no hay a quién actualizarle el costo/precio.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    productos_afectados = list(Producto.objects.filter(
        empresa=empresa_actual, articulo_color_fotos=costeo.referencia_color, activo=True
    ))
    nuevo_costo = costeo.costo_unitario
    nuevo_precio = costeo.precio_venta_unitario

    if request.method == 'POST':
        Producto.objects.filter(
            empresa=empresa_actual, articulo_color_fotos=costeo.referencia_color, activo=True
        ).update(costo=nuevo_costo, precio_venta=nuevo_precio)
        messages.success(
            request,
            f"Costo (${nuevo_costo:,.2f}) y precio de venta (${nuevo_precio:,.2f}) actualizados en "
            f"{len(productos_afectados)} talla(s) de {costeo.referencia_color}."
        )
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    context = {
        'costeo': costeo,
        'productos_afectados': productos_afectados,
        'nuevo_costo': nuevo_costo,
        'nuevo_precio': nuevo_precio,
    }
    return render(request, 'costeo_jeans/costeo_actualizar_catalogo_confirmar.html', context)


@login_required
@_rol_requerido
@require_POST
def costeo_finalizar(request, costeo_id):
    """
    Pasa el costeo a Finalizado: bloquea su edición y descuenta los insumos
    usados del inventario (una sola vez, controlado por
    Costeo.stock_descontado -- ver nota en models.py sobre por qué esto ya
    no es una señal).
    """
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(Costeo.objects.prefetch_related('detalle_insumos__insumo'), id=costeo_id, empresa=empresa_actual)

    if costeo.estado == Costeo.EstadoCosteo.FINALIZADO:
        messages.info(request, "Este costeo ya estaba Finalizado.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    if not costeo.detalle_insumos.exists() and not costeo.detalle_procesos.exists():
        messages.error(request, "Agrega al menos un insumo o proceso antes de finalizar el costeo.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    with transaction.atomic():
        if not costeo.stock_descontado:
            for detalle in costeo.detalle_insumos.all():
                MovimientoInsumo.objects.create(
                    insumo=detalle.insumo,
                    tipo=MovimientoInsumo.Tipo.SALIDA,
                    cantidad=detalle.cantidad,
                    descripcion=f"Uso en producción para costeo: {costeo.referencia}",
                    costeo_relacionado=costeo,
                )
            costeo.stock_descontado = True
        costeo.estado = Costeo.EstadoCosteo.FINALIZADO
        costeo.save(update_fields=['estado', 'stock_descontado'])

    messages.success(request, "Costeo finalizado. Se descontaron los insumos del inventario y el costeo quedó bloqueado para edición.")
    return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)


@login_required
@user_passes_test(es_administracion, login_url='core:acceso_denegado')
@require_POST
def costeo_reabrir(request, costeo_id):
    """
    Solo Administración: regresa el costeo a Borrador y revierte el
    descuento de inventario con movimientos de ENTRADA de reversión (nunca
    se borran movimientos -- mismo criterio append-only que usa
    MovimientoInventario en la app bodega).
    """
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(Costeo, id=costeo_id, empresa=empresa_actual)

    if costeo.estado != Costeo.EstadoCosteo.FINALIZADO:
        messages.info(request, "Este costeo no está Finalizado.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    with transaction.atomic():
        if costeo.stock_descontado:
            salidas = MovimientoInsumo.objects.filter(costeo_relacionado=costeo, tipo=MovimientoInsumo.Tipo.SALIDA)
            for salida in salidas:
                MovimientoInsumo.objects.create(
                    insumo=salida.insumo,
                    tipo=MovimientoInsumo.Tipo.ENTRADA,
                    cantidad=salida.cantidad,
                    descripcion=f"Reversión por reapertura de Costeo #{costeo.pk} ({costeo.referencia})",
                    costeo_relacionado=costeo,
                )
            costeo.stock_descontado = False
        costeo.estado = Costeo.EstadoCosteo.BORRADOR
        costeo.save(update_fields=['estado', 'stock_descontado'])

    messages.success(request, "Costeo reabierto. Los insumos descontados fueron devueltos al inventario.")
    return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)


def _costeos_historial_filtrados(request):
    """Mismo filtro (referencia/fecha) que usa CosteoHistoryListView, reutilizado por la exportación a Excel."""
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Costeo.objects.none()
    queryset = Costeo.objects.filter(empresa=empresa_actual).order_by('-fecha')
    query_ref = request.GET.get('referencia')
    query_date = request.GET.get('fecha')
    if query_ref:
        queryset = queryset.filter(referencia__icontains=query_ref)
    if query_date:
        queryset = queryset.filter(fecha=query_date)
    return queryset


class CosteoHistoryListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = Costeo
    template_name = 'costeo_jeans/costeo_historial.html'
    context_object_name = 'costeos'
    paginate_by = 20
    def get_queryset(self):
        return _costeos_historial_filtrados(self.request)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_ref'] = self.request.GET.get('referencia', '')
        context['search_date'] = self.request.GET.get('fecha', '')
        return context


@login_required
@_rol_requerido
def costeo_historial_exportar_excel(request):
    """Exporta el historial de costeos (con los mismos filtros activos) a Excel."""
    costeos = _costeos_historial_filtrados(request)

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = 'Historial de Costeos'

    encabezados = [
        'Referencia', 'Fecha', 'Estado', 'Cantidad Producida', 'Costo Unitario', 'Costo Total',
        'Margen Deseado %', 'Precio de Venta', 'Utilidad Unitaria', 'Margen Utilidad %',
        'Utilidad Neta Unitaria', 'Margen Neto %',
    ]
    hoja.append(encabezados)
    for indice, encabezado in enumerate(encabezados, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = max(len(encabezado) + 2, 16)

    for costeo in costeos:
        hoja.append([
            costeo.referencia,
            costeo.fecha.strftime('%Y-%m-%d'),
            costeo.get_estado_display(),
            costeo.cantidad_producida,
            float(costeo.costo_unitario),
            float(costeo.costo_total),
            float(costeo.margen_deseado),
            float(costeo.precio_venta_unitario),
            float(costeo.utilidad_unitaria),
            float(costeo.margen_utilidad),
            float(costeo.utilidad_neta_unitaria),
            float(costeo.margen_neto),
        ])

    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_costeos.xlsx"'
    return response


class InformesView(LoginRequiredMixin, CosteoRolMixin, TemplateView):
    template_name = 'costeo_jeans/informes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_actual = getattr(self.request, 'tenant', None)

        costeos = Costeo.objects.filter(empresa=empresa_actual)

        if not costeos.exists():
            context['no_data'] = True
            return context

        # --- 1. KPIs Generales ---
        # Costo unitario promedio ponderado: Sum(costo_total)/Sum(unidades) --
        # NO Avg(costo_total)/Avg(unidades) (eso sería un promedio de
        # promedios, sesgado si los costeos tienen tamaños de producción
        # muy distintos).
        kpis = costeos.aggregate(
            costo_total_general=Sum('costo_total'),
            unidades_totales=Sum('cantidad_producida'),
        )
        kpis['costo_unitario_promedio'] = (
            kpis['costo_total_general'] / kpis['unidades_totales']
            if kpis['unidades_totales'] else 0
        )
        context['kpis'] = kpis

        # --- 1b. KPIs de precio de venta y utilidad (propiedades Python,
        # no se pueden agregar en SQL -- se calculan iterando una sola vez). ---
        costeos_lista = list(costeos)
        precios = [c.precio_venta_unitario for c in costeos_lista if c.precio_venta_unitario]
        margenes = [c.margen_utilidad for c in costeos_lista if c.precio_venta_unitario]
        context['precio_venta_promedio'] = sum(precios) / len(precios) if precios else 0
        context['margen_promedio'] = sum(margenes) / len(margenes) if margenes else 0
        context['utilidad_total_estimada'] = sum(c.utilidad_total for c in costeos_lista)

        # --- 2. NUEVOS KPIs ADICIONALES ---
        # Conteo de referencias de producto únicas
        context['conteo_referencias'] = costeos.values('referencia').distinct().count()

        # Referencia más producida (agrupando por referencia y sumando la cantidad)
        ref_mas_producida = costeos.values('referencia').annotate(
            total=Sum('cantidad_producida')
        ).order_by('-total').first()
        context['ref_mas_producida'] = ref_mas_producida


        # --- 3. Datos para Gráfico de Distribución de Costos (sin cambios) ---
        total_insumos = sum(d.costo_total for d in DetalleInsumo.objects.filter(costeo__in=costeos))
        total_procesos = sum(d.costo_total for d in DetalleProceso.objects.filter(costeo__in=costeos))
        total_fijos = sum(d.costo_total for d in DetalleCostoFijo.objects.filter(costeo__in=costeos))

        context['distribucion_costos_data'] = json.dumps({
            "labels": ["Insumos", "Procesos", "Costos Fijos/Generales"],
            "data": [float(total_insumos), float(total_procesos), float(total_fijos)],
        })

        # --- 4. Datos para Gráfico de Evolución de Costos Unitarios ---
        # Igual que el KPI general: promedio ponderado por mes (Sum/Sum), no
        # promedio de promedios.
        evolucion_costos = list(costeos.annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            costo_total_mes=Sum('costo_total'),
            unidades_mes=Sum('cantidad_producida'),
        ).order_by('mes'))

        for e in evolucion_costos:
            e['costo_promedio'] = e['costo_total_mes'] / e['unidades_mes'] if e['unidades_mes'] else None

        evolucion_data = {
            "labels": [e['mes'].strftime('%b %Y') for e in evolucion_costos],
            "data": [round(float(e['costo_promedio']), 2) for e in evolucion_costos if e['costo_promedio'] is not None],
        }
        context['evolucion_costos_data'] = json.dumps(evolucion_data, cls=DjangoJSONEncoder)

        # --- 5. Top 5 Referencias más costosas ---
        costeos_validos = [c for c in costeos_lista if c.cantidad_producida > 0]
        top_costosos = sorted(costeos_validos, key=lambda c: c.costo_unitario, reverse=True)[:5]
        context['top_costosos'] = top_costosos

        context['titulo'] = "Informes y Estadísticas"
        return context
    
    
@login_required
@_rol_requerido
def export_costeo_pdf(request, costeo_id):
    """
    Genera y descarga un PDF para un costeo específico.
    """
    empresa_actual = getattr(request, 'tenant', None)

    # Obtenemos el objeto costeo con todos sus detalles (igual que en la vista de resumen)
    costeo = get_object_or_404(
        Costeo.objects.prefetch_related(
            'detalle_insumos__insumo',
            'detalle_tallas',
            'detalle_procesos__tarifa__proceso',
            'detalle_procesos__tarifa__confeccionista',
            'detalle_costos_fijos__costo_fijo'
        ),
        id=costeo_id,
        empresa=empresa_actual
    )

    # Renderizamos la plantilla de resumen a un string de HTML
    html_string = render_to_string('costeo_jeans/costeo_pdf_template.html', {'costeo': costeo})

    # Creamos el PDF en memoria
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    # Creamos la respuesta HTTP para la descarga
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="costeo-{costeo.referencia}-{costeo.id}.pdf"'

    return response


@login_required
@_rol_requerido
def costeo_update_step1(request, costeo_id):
    """
    Paso 1 para EDITAR un costeo existente (datos principales + desglose por talla).
    """
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(Costeo, id=costeo_id, empresa=empresa_actual)

    if costeo.estado == Costeo.EstadoCosteo.FINALIZADO:
        messages.error(request, "Este costeo está Finalizado y no se puede editar. Reábrelo primero si necesitas ajustarlo.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    if request.method == 'POST':
        # Pasamos 'instance=costeo' para que el formulario actualice el objeto existente
        form = CosteoModelForm(request.POST, instance=costeo, empresa=empresa_actual)
        talla_formset = DetalleTallaFormSet(request.POST, instance=costeo, prefix='tallas')
        if form.is_valid() and talla_formset.is_valid():
            with transaction.atomic():
                form.save()
                talla_formset.save()
                nueva_cantidad = costeo.detalle_tallas.aggregate(total=Sum('cantidad'))['total'] or 0
                if nueva_cantidad != costeo.cantidad_producida:
                    costeo.cantidad_producida = nueva_cantidad
                    costeo.save(update_fields=['cantidad_producida'])
                    # Los insumos ya cargados calculan su 'cantidad' a partir
                    # de costeo.cantidad_producida -- si cambió, hay que
                    # re-guardarlos para que no queden con un total viejo.
                    for detalle in costeo.detalle_insumos.all():
                        detalle.save()
            return redirect('costeo_jeans:costeo_update_step2', costeo_id=costeo.id)
    else:
        # Pasamos 'instance=costeo' para que el formulario se cargue con los datos existentes
        form = CosteoModelForm(instance=costeo, empresa=empresa_actual)
        talla_formset = DetalleTallaFormSet(instance=costeo, prefix='tallas')

    return render(request, 'costeo_jeans/costeo_create_step1.html', {
        'form': form,
        'talla_formset': talla_formset,
        'titulo': f'Editando Costeo: {costeo.referencia}' # Título personalizado para la edición
    })


@login_required
@_rol_requerido
def costeo_update_step2(request, costeo_id):
    empresa_actual = getattr(request, 'tenant', None)
    costeo = get_object_or_404(Costeo, id=costeo_id, empresa=empresa_actual)

    if costeo.estado == Costeo.EstadoCosteo.FINALIZADO:
        messages.error(request, "Este costeo está Finalizado y no se puede editar. Reábrelo primero si necesitas ajustarlo.")
        return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    insumos = Insumo.objects.filter(empresa=empresa_actual)
    insumo_categorias = {insumo.id: insumo.categoria for insumo in insumos}

    insumo_formset = DetalleInsumoFormSet(
        request.POST or None,
        instance=costeo,
        prefix='insumos',
        form_kwargs={'empresa': empresa_actual}
    )
    proceso_formset = DetalleProcesoFormSet(
        request.POST or None,
        instance=costeo,
        prefix='procesos',
        form_kwargs={'empresa': empresa_actual}
    )

    if request.method == 'POST':
        is_insumos_valid = insumo_formset.is_valid()
        is_procesos_valid = proceso_formset.is_valid()

        if is_insumos_valid and is_procesos_valid:
            with transaction.atomic():
                insumo_formset.save()
                proceso_formset.save()
            return redirect('costeo_jeans:costeo_summary', costeo_id=costeo.id)

    context = {
        'costeo': costeo,
        'insumo_formset': insumo_formset,
        'proceso_formset': proceso_formset,
        'insumo_categorias_json': json.dumps(insumo_categorias)
    }
    return render(request, 'costeo_jeans/costeo_create_step2.html', context)

class TarifaListView(LoginRequiredMixin, CosteoRolMixin, ListView):
    model = TarifaConfeccionista
    template_name = 'costeo_jeans/tarifa_list.html'
    context_object_name = 'tarifas'
    paginate_by = 10

    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return TarifaConfeccionista.objects.filter(empresa=empresa_actual).select_related('confeccionista', 'proceso').order_by('confeccionista__nombre')
        return TarifaConfeccionista.objects.none()

class TarifaCreateView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = TarifaConfeccionista
    form_class = TarifaConfeccionistaForm
    template_name = 'costeo_jeans/tarifa_form.html'
    success_url = reverse_lazy('costeo_jeans:tarifa_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = getattr(self.request, 'tenant', None)
        return kwargs

    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'tenant', None)
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Añadir Nueva Tarifa'
        context['boton'] = 'Guardar Tarifa'
        return context

class TarifaUpdateView(LoginRequiredMixin, CosteoRolMixin, UpdateView):
    model = TarifaConfeccionista
    form_class = TarifaConfeccionistaForm
    template_name = 'costeo_jeans/tarifa_form.html'
    success_url = reverse_lazy('costeo_jeans:tarifa_list')

    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return TarifaConfeccionista.objects.filter(empresa=empresa_actual)
        return TarifaConfeccionista.objects.none()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = getattr(self.request, 'tenant', None)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Tarifa'
        context['boton'] = 'Actualizar Tarifa'
        return context

class TarifaDeleteView(LoginRequiredMixin, CosteoRolMixin, DeleteView):
    model = TarifaConfeccionista
    template_name = 'costeo_jeans/tarifa_confirm_delete.html'
    success_url = reverse_lazy('costeo_jeans:tarifa_list')

    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        if empresa_actual:
            return TarifaConfeccionista.objects.filter(empresa=empresa_actual)
        return TarifaConfeccionista.objects.none()
    
class InsumoDetailView(LoginRequiredMixin, CosteoRolMixin, DetailView):
    model = Insumo
    template_name = 'costeo_jeans/insumo_detail.html'
    context_object_name = 'insumo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movimientos'] = MovimientoInsumo.objects.filter(insumo=self.object)
        context['titulo'] = f"Historial de Inventario: {self.object.nombre}"
        return context

# --- NUEVA VISTA: Para registrar entradas (compras) ---
class RegistrarEntradaInsumoView(LoginRequiredMixin, CosteoRolMixin, CreateView):
    model = MovimientoInsumo
    form_class = MovimientoInsumoForm
    template_name = 'costeo_jeans/movimiento_insumo_form.html'

    def get_success_url(self):
        return reverse_lazy('costeo_jeans:insumo_detail', kwargs={'pk': self.object.insumo.pk})

    def form_valid(self, form):
        form.instance.tipo = MovimientoInsumo.Tipo.ENTRADA
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = "Registrar Entrada de Insumo"
        return context

# --- NUEVA VISTA: Para generar el PDF de movimientos ---
@login_required
@_rol_requerido
def export_movimientos_pdf(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk, empresa=getattr(request, 'tenant', None))
    movimientos = MovimientoInsumo.objects.filter(insumo=insumo)

    html_string = render_to_string('costeo_jeans/movimiento_insumo_pdf.html', {
        'insumo': insumo,
        'movimientos': movimientos
    })

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historial-{insumo.nombre}.pdf"'
    return response