from collections import defaultdict
from decimal import Decimal
from core.auth_utils import es_administracion
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse
from django.views.generic import UpdateView
from django.views.generic import ListView
from django.views.generic import CreateView
from django.views.generic import DeleteView
from django.urls import reverse_lazy
from bodega.models import MovimientoInventario, CabeceraConteo, ConteoInventario, Bodega, AccesoBodega, TrasladoBodega, DetalleTrasladoBodega
from django.shortcuts import render, get_object_or_404, redirect
from pedidos.models import Pedido
from django.http import Http404, JsonResponse
import openpyxl
import csv
from django.views.generic import DetailView
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction, IntegrityError
from .models import BorradorDespacho, DetalleBorradorDespacho, SalidaInternaCabecera
from .forms import DetalleIngresoModificarFormSet, SalidaInternaCabeceraForm, DetalleSalidaInternaFormSet
import json
from django.utils.decorators import method_decorator
from django.contrib.auth import get_user_model
import logging
from django.conf import settings
from django.db.models import Prefetch, Value
from django.db.models.functions import Coalesce
from itertools import groupby
from pedidos.models import DetallePedido
from django.db.models import Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from core.utils import render_pdf_weasyprint, get_logo_base_64_despacho
from weasyprint import HTML
from bodega.forms import DetalleIngresoFormSet, InfoGeneralConteoForm, IngresoBodegaForm
from .forms import InfoGeneralConteoForm, ImportarConteoForm
from productos.models import Producto
from django.template.loader import render_to_string
from .models import IngresoBodega
from .models import ComprobanteDespacho, DetalleComprobanteDespacho
from django.db.models import Max, F
from core.mixins import TenantAwareMixin
from .resources import PlantillaConteoResource
from factura.models import EstadoFacturaDespacho
from .models import CambioProducto
from django.db import transaction
from .forms import CambioProductoForm
from .forms import BodegaForm, AccesoBodegaForm, TrasladoBodegaForm
from django.utils.safestring import mark_safe
from django.core.paginator import Paginator
from django.db.models import Q
from django.views import View


def admin_permission_required(view_func):
    """
    Decorador que verifica si el usuario es superusuario o pertenece al grupo 'administracion'.
    """
    @login_required(login_url='core:acceso_denegado')
    def _wrapped_view_func(request, *args, **kwargs):
        user = request.user
        if user.is_superuser or user.groups.filter(name='administracion').exists():
            return view_func(request, *args, **kwargs)
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('core:acceso_denegado')
    return _wrapped_view_func




@login_required
@permission_required('pedidos.change_pedido', login_url='core:acceso_denegado')
def vista_despacho_pedido(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.") 
        return redirect('core:index')
    
    # --- INICIO: Cargar Mapeo de Tallas (Refactor) ---
    TALLAS_MAPEO = empresa_actual.talla_mapeo or {}
    # --- FIN: Cargar Mapeo de Tallas ---
    
    pedido = get_object_or_404(
        Pedido.objects.prefetch_related('detalles__producto'), 
        pk=pk, 
        empresa=empresa_actual
    )
    
    ESTADOS_PERMITIDOS_PARA_DESPACHO = ['APROBADO_ADMIN', 'PROCESANDO', 'PENDIENTE_BODEGA', 'LISTO_BODEGA_DIRECTO']

    if request.method == 'POST':
        # ... (Tu lógica POST se mantiene igual, no es necesario cambiarla) ...
        pass # Dejo esto como referencia, no borres tu lógica POST

    if pedido.estado not in ESTADOS_PERMITIDOS_PARA_DESPACHO and not request.user.is_superuser:
        messages.info(request, f"El pedido #{pedido.numero_pedido_empresa} ya está en estado '{pedido.get_estado_display()}' y no puede ser modificado aquí.")
        return redirect('bodega:lista_pedidos_bodega')
    
    # --- INICIA EL BLOQUE CORREGIDO ---

    cantidades_en_borrador_map = {} # Inicializamos la variable aquí para que siempre exista

    with transaction.atomic():
        pedido_locked = Pedido.objects.select_for_update().get(pk=pk)

        borrador, created = BorradorDespacho.objects.get_or_create(
            pedido=pedido_locked,
            empresa=empresa_actual,
            defaults={'usuario': request.user}
        )

        if created:
            detalles_para_borrador = []
            for detalle in pedido_locked.detalles.all():
                if detalle.cantidad_verificada and detalle.cantidad_verificada > 0:
                    detalles_para_borrador.append(
                        DetalleBorradorDespacho(
                            borrador_despacho=borrador,
                            detalle_pedido_origen=detalle,
                            producto=detalle.producto,
                            cantidad_escaneada_en_borrador=detalle.cantidad_verificada
                        )
                    )
            if detalles_para_borrador:
                DetalleBorradorDespacho.objects.bulk_create(detalles_para_borrador)
                logger.info(f"Nuevo BorradorDespacho {borrador.pk} poblado con datos de Pedido {pk}.")

        # Llenamos el diccionario desde el borrador, que ahora está garantizado que existe y está sincronizado
        cantidades_en_borrador_map = {
            db.detalle_pedido_origen_id: db.cantidad_escaneada_en_borrador 
            for db in borrador.detalles_borrador.all()
        }
    
    # --- FIN DEL BLOQUE CORREGIDO ---

    detalles_pedido = pedido.detalles.all().order_by('producto__referencia', 'producto__talla')

    items_agrupados_dict = defaultdict(lambda: {
        'referencia': '', 'nombre': '', 'color': '', 'tallas': [],
        'total_pedida': 0, 'total_verificada': 0
    })

    for detalle in detalles_pedido:
        producto_obj = detalle.producto
        clave_agrupacion = (producto_obj.referencia, producto_obj.color or '')
        
        cantidad_a_mostrar_en_frontend = cantidades_en_borrador_map.get(detalle.pk, 0)

        if not items_agrupados_dict[clave_agrupacion]['referencia']:
            items_agrupados_dict[clave_agrupacion]['referencia'] = producto_obj.referencia
            items_agrupados_dict[clave_agrupacion]['nombre'] = producto_obj.nombre
            items_agrupados_dict[clave_agrupacion]['color'] = producto_obj.color or ''

# --- Aplicar Mapeo de Talla (Refactor) ---
        talla_original = producto_obj.talla or ''
        talla_como_texto = str(talla_original).strip()
        talla_display = TALLAS_MAPEO.get(talla_como_texto, talla_como_texto)
        # --- Fin Mapeo ---
        
        items_agrupados_dict[clave_agrupacion]['tallas'].append({
            'nombre': talla_display, # <-- LÍNEA MODIFICADA
            'detalle': {
                'id': detalle.pk,
                'codigo_barras': producto_obj.codigo_barras or '', 
                'cantidad_pedida': detalle.cantidad,
                'cantidad_verificada': cantidad_a_mostrar_en_frontend,
            }
        })
        items_agrupados_dict[clave_agrupacion]['total_pedida'] += detalle.cantidad
        items_agrupados_dict[clave_agrupacion]['total_verificada'] += cantidad_a_mostrar_en_frontend

    detalles_json_data = list(items_agrupados_dict.values())
    detalles_json_data.sort(key=lambda x: (x['referencia'], x['color']))
    detalles_json = json.dumps(detalles_json_data) 
    
    context = {
        'pedido': pedido,
        'detalles_pedido': detalles_pedido,
        'detalles_json': detalles_json,
        'titulo': f"Despacho Pedido #{pedido.numero_pedido_empresa}",
    }
    return render(request, 'bodega/despacho_pedido.html', context)

@csrf_exempt
@require_POST
@login_required
@permission_required('pedidos.change_pedido', login_url='core:acceso_denegado')
def guardar_parcialmente_detalle_ajax(request, pk): # 'pk' es el pedido_id
    """
    Guarda el progreso de la verificación (cantidades escaneadas en borrador) de uno o más DetallePedido
    en el modelo BorradorDespacho. No afecta DetallePedido.cantidad_verificada ni genera comprobantes.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        logger.error("Error: Empresa no identificada en guardar_parcialmente_detalle_ajax.")
        return JsonResponse({'status': 'error', 'message': 'Empresa no identificada.'}, status=400)

    try:
        data = json.loads(request.body)
        detalles_a_guardar_frontend = data.get('detalles_a_guardar', [])

        if not detalles_a_guardar_frontend:
            return JsonResponse({'status': 'info', 'message': 'No se enviaron detalles para guardar.'}, status=200)

        with transaction.atomic():
            pedido = get_object_or_404(Pedido, pk=pk, empresa=empresa_actual)

            # Obtener o crear el borrador para este pedido y usuario
            borrador, _ = BorradorDespacho.objects.get_or_create(
                pedido=pedido,
                empresa=empresa_actual,
                defaults={'usuario': request.user}
            )

            # Mapear los detalles del pedido original y los del borrador existente
            detalles_pedido_map = {d.pk: d for d in pedido.detalles.select_for_update()} # Bloquear los DetallePedido
            detalles_borrador_map = {
                db.detalle_pedido_origen_id: db 
                for db in borrador.detalles_borrador.select_for_update() # Bloquear los DetalleBorradorDespacho
            }

            cambios_realizados = 0
            for item_data in detalles_a_guardar_frontend:
                detalle_id = item_data.get('id')
                cantidad_escaneada_borrador = item_data.get('cantidad_verificada_nueva') # Es la cantidad de escaneos para el borrador

                if detalle_id is None or cantidad_escaneada_borrador is None:
                    logger.warning(f"DEBUG: Detalle con datos incompletos en el payload: {item_data}")
                    continue

                try:
                    cantidad_escaneada_borrador = int(cantidad_escaneada_borrador)
                    if cantidad_escaneada_borrador < 0:
                        logger.warning(f"DEBUG: Cantidad negativa para detalle {detalle_id}: {cantidad_escaneada_borrador}")
                        continue
                except ValueError:
                    logger.warning(f"DEBUG: Cantidad no válida para detalle {detalle_id}: {cantidad_escaneada_borrador}")
                    continue

                detalle_original = detalles_pedido_map.get(detalle_id)
                if not detalle_original:
                    logger.warning(f"DEBUG: Detalle {detalle_id} no encontrado o no pertenece a este pedido/empresa.")
                    continue

                # Validar que la cantidad escaneada en borrador no exceda la pedida original
                if cantidad_escaneada_borrador > detalle_original.cantidad:
                    logger.warning(f"DEBUG: Exceso de cantidad para {detalle_original.producto.referencia}: {cantidad_escaneada_borrador} > {detalle_original.cantidad}.")
                    # Aquí puedes decidir si lanzar un error fatal o simplemente ignorar el exceso
                    cantidad_escaneada_borrador = detalle_original.cantidad # Limitar al máximo pedido

                detalle_borrador_obj = detalles_borrador_map.get(detalle_id)

                if detalle_borrador_obj: # Si ya existe un detalle de borrador para este DetallePedido
                    if detalle_borrador_obj.cantidad_escaneada_en_borrador != cantidad_escaneada_borrador:
                        detalle_borrador_obj.cantidad_escaneada_en_borrador = cantidad_escaneada_borrador
                        detalle_borrador_obj.save(update_fields=['cantidad_escaneada_en_borrador'])
                        cambios_realizados += 1
                        logger.info(f"DetalleBorradorDespacho {detalle_borrador_obj.pk} actualizado. Cant. Borrador: {cantidad_escaneada_borrador}.")
                else: # Si no existe, crear uno nuevo
                    DetalleBorradorDespacho.objects.create(
                        borrador_despacho=borrador,
                        detalle_pedido_origen=detalle_original,
                        producto=detalle_original.producto, # El producto del detalle original
                        cantidad_escaneada_en_borrador=cantidad_escaneada_borrador
                    )
                    cambios_realizados += 1
                    logger.info(f"Nuevo DetalleBorradorDespacho creado para {detalle_original.pk}. Cant. Borrador: {cantidad_escaneada_borrador}.")

            # Actualizar el estado del pedido a 'PROCESANDO' si se guarda progreso
            if cambios_realizados > 0 and pedido.estado in ['APROBADO_ADMIN', 'PROCESANDO', 'PENDIENTE_BODEGA', 'LISTO_BODEGA_DIRECTO']:
                pedido.estado = 'PROCESANDO'
                pedido.save(update_fields=['estado'])
                logger.info(f"Pedido {pedido.pk} estado cambiado a PROCESANDO.")

            if cambios_realizados > 0:
                return JsonResponse({'status': 'success', 'message': f'Progreso de {cambios_realizados} ítems guardado exitosamente.'})
            else:
                return JsonResponse({'status': 'info', 'message': 'No se detectaron cambios significativos para guardar.'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Formato JSON inválido.'}, status=400)
    except Exception as e:
        logger.error(f"Error inesperado en guardar_parcialmente_detalle_ajax para pedido {pk}: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Error interno del servidor: {str(e)}'}, status=500)
    
    
@csrf_exempt
@require_POST
@login_required
@permission_required('pedidos.change_pedido', login_url='core:acceso_denegado')
def enviar_despacho_parcial_ajax(request, pk):
    """
    Procesa un envío del despacho, tomando las cantidades del BorradorDespacho,
    generando un NUEVO ComprobanteDespacho, actualizando DetallePedido.cantidad_verificada
    y eliminando el borrador.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        logger.error("Error: Empresa no identificada en enviar_despacho_parcial_ajax.")
        return JsonResponse({'status': 'error', 'message': 'Empresa no identificada.'}, status=400)

    try:
        with transaction.atomic():
            pedido = get_object_or_404(
                Pedido.objects.select_for_update().prefetch_related('detalles'), 
                pk=pk, 
                empresa=empresa_actual
            )

            ESTADOS_PERMITIDOS_ENVIO = ['APROBADO_ADMIN', 'PROCESANDO', 'PENDIENTE_BODEGA', 'LISTO_BODEGA_DIRECTO']
            if pedido.estado not in ESTADOS_PERMITIDOS_ENVIO and not request.user.is_superuser:
                raise ValueError(f"El pedido #{pedido.numero_pedido_empresa} no puede ser enviado en su estado actual ({pedido.get_estado_display()}).")

            borrador = BorradorDespacho.objects.filter(pedido=pedido, empresa=empresa_actual).first()

            if not borrador or not borrador.detalles_borrador.exists():
                return JsonResponse({'status': 'warning', 'message': 'No hay unidades escaneadas en borrador para enviar.'}, status=200)

            items_para_este_comprobante = []
            detalles_pedido_actualizados = []

            for detalle_borrador in borrador.detalles_borrador.select_for_update():
                detalle_original = detalle_borrador.detalle_pedido_origen
                cantidad_a_comprobar_en_esta_op = detalle_borrador.cantidad_escaneada_en_borrador - (detalle_original.cantidad_verificada or 0)

                if cantidad_a_comprobar_en_esta_op > 0:
                    items_para_este_comprobante.append({
                        'producto': detalle_original.producto,
                        'cantidad_despachada': cantidad_a_comprobar_en_esta_op,
                        'detalle_pedido_origen': detalle_original
                    })

                    detalle_original.cantidad_verificada = detalle_borrador.cantidad_escaneada_en_borrador
                    detalle_original.verificado_bodega = True
                    detalles_pedido_actualizados.append(detalle_original)

            if not items_para_este_comprobante:
                return JsonResponse({'status': 'warning', 'message': 'No se detectaron nuevas unidades para despachar en el borrador.'}, status=200)

            # --- LÓGICA CLAVE: Siempre se crea un NUEVO comprobante ---
            comprobante = ComprobanteDespacho.objects.create(
                pedido=pedido,
                empresa=empresa_actual,
                fecha_hora_despacho=timezone.now(),
                usuario_responsable=request.user,
            )
            logger.info(f"NUEVO Comprobante {comprobante.pk} creado.")

            detalles_comprobante_bulk_create = [
                DetalleComprobanteDespacho(
                    comprobante_despacho=comprobante,
                    producto=item_data['producto'],
                    cantidad_despachada=item_data['cantidad_despachada'],
                    detalle_pedido_origen=item_data['detalle_pedido_origen']
                ) for item_data in items_para_este_comprobante
            ]
            DetalleComprobanteDespacho.objects.bulk_create(detalles_comprobante_bulk_create)

            if detalles_pedido_actualizados:
                DetallePedido.objects.bulk_update(detalles_pedido_actualizados, ['cantidad_verificada', 'verificado_bodega'])

            borrador.delete()
            logger.info(f"Borrador de despacho {borrador.pk} eliminado tras generar comprobante.")

            # Actualizar estado del pedido
            pedido.refresh_from_db()
            todos_los_items_despachados_completamente = all(
                (d.cantidad_verificada or 0) >= d.cantidad for d in pedido.detalles.all()
            )

            comprobante.es_parcial = not todos_los_items_despachados_completamente
            comprobante.save(update_fields=['es_parcial'])

            if todos_los_items_despachados_completamente:
                pedido.estado = 'ENVIADO'
            else:
                pedido.estado = 'PROCESANDO'
            pedido.save(update_fields=['estado'])
            
            EstadoFacturaDespacho.objects.create(empresa=empresa_actual, despacho=comprobante)
            comprobante_url = reverse('bodega:imprimir_comprobante_especifico', kwargs={'pk_comprobante': comprobante.pk})

            return JsonResponse({'status': 'success', 'message': f'Nuevo comprobante #{comprobante.pk} generado.', 'comprobante_url': comprobante_url})

    except Exception as e:
        logger.critical(f"Error inesperado en enviar_despacho_parcial_ajax para pedido {pk}: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Error interno del servidor: {str(e)}'}, status=500)

@login_required
@permission_required('bodega.view_lista_pedidos_bodega', login_url='core:acceso_denegado')
def vista_lista_pedidos_bodega(request):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')

    # --- 1. Obtener parámetros de búsqueda (NUEVO FORMATO UNIFICADO) ---
    q_query = request.GET.get('q', '').strip()
    estado_query = request.GET.get('estado', '').strip()

    # --- Preparar Prefetch (sin cambios) ---
    prefetch_detalles = Prefetch(
        'detalles',
        queryset=DetallePedido.objects.select_related('producto'),
        to_attr='detalles_precargados'
    )

    # --- Query base SIN filtro de estado inicial ---
    # NOTA: Agregué 'cliente_online' al select_related para que la búsqueda también encuentre los de Shopify
    pedidos_list = Pedido.objects.filter(
        empresa=empresa_actual
    ).select_related('cliente', 'cliente_online', 'vendedor__user').prefetch_related(prefetch_detalles)

    # --- Determinar qué estados mostrar ---
    estados_validos = [choice[0] for choice in Pedido.ESTADO_PEDIDO_CHOICES] 
    estados_por_defecto = ['APROBADO_ADMIN', 'PROCESANDO', 'LISTO_BODEGA_DIRECTO'] 

    titulo = f'Pedidos Bodega ({empresa_actual.nombre})'
    estado_display_filtro = "Todos (por defecto)"

    if estado_query:
        if estado_query in estados_validos:
            pedidos_list = pedidos_list.filter(estado=estado_query)
            estado_display_filtro = dict(Pedido.ESTADO_PEDIDO_CHOICES).get(estado_query, estado_query)
            titulo = f'Pedidos: {estado_display_filtro} ({empresa_actual.nombre})'
        else:
            messages.warning(request, f"El estado '{estado_query}' no es válido.")
            pedidos_list = Pedido.objects.none()
            titulo = f'Pedidos Bodega (Estado Inválido) - {empresa_actual.nombre}'
    else:
        pedidos_list = pedidos_list.filter(estado__in=estados_por_defecto)
        titulo = f'Pedidos Pendientes Bodega ({empresa_actual.nombre})'
        
    # --- 2. Aplicar el Filtro Mágico Unificado (LA CORRECCIÓN ESTÁ AQUÍ) ---
    if q_query:
        # Preparamos las búsquedas de texto (Nombre cliente estandar, online, vendedor y referencia de producto)
        filtros = (
            Q(cliente__nombre_completo__icontains=q_query) |
            Q(cliente_online__nombre_completo__icontains=q_query) |
            Q(vendedor__user__username__icontains=q_query) |
            Q(detalles__producto__referencia__icontains=q_query)
        )

        # Si el usuario escribió un número, le agregamos la orden de buscar ese consecutivo EXACTO
        if q_query.isdigit():
            filtros |= Q(numero_pedido_empresa=int(q_query))
            # Opcional: También busca en cédulas si digitan un número largo
            filtros |= Q(cliente__identificacion__icontains=q_query)
            filtros |= Q(cliente_online__identificacion__icontains=q_query)

        # Aplicamos todos los filtros de una vez y usamos distinct() para que no salgan duplicados
        pedidos_list = pedidos_list.filter(filtros).distinct()

    # --- Orden final (sin cambios) ---
    pedidos_list = pedidos_list.order_by('-fecha_hora')

    context = {
        'pedidos_list': pedidos_list,
        'titulo': titulo,
        'q_query': q_query, # Actualizado para coincidir con el HTML
        'estado_query': estado_query, 
        'ESTADO_PEDIDO_CHOICES': Pedido.ESTADO_PEDIDO_CHOICES, 
    }
    return render(request, 'bodega/lista_pedidos_bodega.html', context)

@login_required
@permission_required(['pedidos.change_pedido', 'bodega.add_comprobantedespacho'], login_url='core:acceso_denegado')
def vista_verificar_pedido(request, pk):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    pedido = get_object_or_404(Pedido, pk=pk, empresa=empresa_actual)

    if request.method == 'GET':
        # --- INICIO: Cargar Mapeo y Procesar Tallas (Refactor) ---
        TALLAS_MAPEO = empresa_actual.talla_mapeo or {}
        
        # Convertimos a lista para poder modificar las tallas
        detalles_para_mostrar = list(
            pedido.detalles.select_related('producto').all().order_by('producto__referencia', 'producto__color', 'producto__talla')
        )
        
        # --- Separar pendientes de ya completados (pendiente total = 0) ---
        detalles_pendientes = []
        detalles_completados = []
        for detalle in detalles_para_mostrar:
            talla_original = detalle.producto.talla or ''
            talla_como_texto = str(talla_original).strip()
            # Modificamos el atributo 'talla' del objeto producto en memoria
            detalle.producto.talla = TALLAS_MAPEO.get(talla_como_texto, talla_como_texto)

            detalle.cantidad_ya_despachada = detalle.cantidad_verificada or 0
            detalle.cantidad_aun_pendiente = detalle.cantidad - detalle.cantidad_ya_despachada

            if detalle.cantidad_aun_pendiente > 0:
                detalles_pendientes.append(detalle)
            else:
                detalles_completados.append(detalle)
        # --- FIN: Cargar Mapeo y Procesar Tallas ---

        context = {
            'pedido': pedido,
            'detalles_pendientes': detalles_pendientes,
            'detalles_completados': detalles_completados,
            'titulo': f'Verificar pedido número {pedido.numero_pedido_empresa}'
        }
        return render(request, 'bodega/verificar_pedido.html', context)

    elif request.method == 'POST':
        # --- INICIO DE LA VERIFICACIÓN DE CONTRASEÑA ---
        contraseña_ingresada = request.POST.get('contraseña_verificacion_pedido') # Nombre del input en tu plantilla
        contraseña_correcta = settings.CONTRASEÑA_PARA_VERIFICAR_PEDIDO

        if not contraseña_ingresada or contraseña_ingresada != contraseña_correcta:
            messages.error(request, "Contraseña de verificación incorrecta. No se procesaron los cambios.")
            # Redirigir a la misma página (GET) para mostrar el error y el formulario de nuevo
            return redirect('bodega:verificar_pedido', pk=pedido.pk)
        print(f"--- VISTA_VERIFICAR_PEDIDO (POST): INICIO para Pedido #{pk} ---")

        ESTADOS_PERMITIDOS_VERIFICACION = ['APROBADO_ADMIN', 'PENDIENTE', 'PENDIENTE_BODEGA', 'PROCESANDO', 'LISTO_BODEGA_DIRECTO']
        if pedido.estado not in ESTADOS_PERMITIDOS_VERIFICACION and not request.user.is_superuser : # Superusuario puede forzar
            messages.error(request, f"El pedido #{pedido.numero_pedido_empresa} no se puede modificar en su estado actual ({pedido.get_estado_display()}).")
            return redirect('bodega:verificar_pedido', pk=pedido.pk)

        items_efectivamente_despachados_para_comprobante = [] 
        hubo_cambios_generales_en_detalle_pedido = False 

        try:
            with transaction.atomic(): 
                print("Dentro de transaction.atomic(). Iniciando bucle for detalles...")
                detalles_a_procesar = pedido.detalles.select_related('producto').all().order_by('pk')

                for detalle_pedido_origen in detalles_a_procesar: 
                    input_name = f'cantidad_a_despachar_{detalle_pedido_origen.id}'
                    cantidad_a_despachar_str = request.POST.get(input_name)
                    print(f"  - DetallePedido ID {detalle_pedido_origen.id} ({detalle_pedido_origen.producto}): Leyendo '{input_name}', Valor recibido='{cantidad_a_despachar_str}'")

                    cantidad_ya_despachada_total_en_pedido = detalle_pedido_origen.cantidad_verificada or 0
                    cantidad_a_despachar_ahora_int = 0

                    try:
                        if cantidad_a_despachar_str is not None and cantidad_a_despachar_str.strip().isdigit():
                            cantidad_a_despachar_ahora_int = int(cantidad_a_despachar_str)
                        elif cantidad_a_despachar_str is not None and cantidad_a_despachar_str.strip() != '':
                            raise ValueError(f"Valor no numérico '{cantidad_a_despachar_str}' ingresado.")

                        if cantidad_a_despachar_ahora_int < 0:
                            raise ValueError("Cantidad a despachar no puede ser negativa.")

                        
                        pendiente_total_item_pedido = detalle_pedido_origen.cantidad - cantidad_ya_despachada_total_en_pedido
                        if cantidad_a_despachar_ahora_int > pendiente_total_item_pedido:
                            error_msg = (
                                f"Intenta despachar {cantidad_a_despachar_ahora_int} de {detalle_pedido_origen.producto.nombre} "
                                f"pero solo quedan {pendiente_total_item_pedido} pendientes en el pedido."
                            )
                            messages.error(request, error_msg)
                            raise transaction.TransactionManagementError(error_msg)

                        
                        if cantidad_a_despachar_ahora_int > 0:
                            items_efectivamente_despachados_para_comprobante.append({
                                'producto': detalle_pedido_origen.producto, 
                                'cantidad_despachada': cantidad_a_despachar_ahora_int,
                                'detalle_pedido_origen': detalle_pedido_origen 
                            })
                            print(f"      Agregado para Comprobante: {detalle_pedido_origen.producto.nombre} - Cantidad: {cantidad_a_despachar_ahora_int}")

                        
                        nuevo_total_verificado_pedido = cantidad_ya_despachada_total_en_pedido + cantidad_a_despachar_ahora_int
                        if nuevo_total_verificado_pedido != cantidad_ya_despachada_total_en_pedido or \
                           (not detalle_pedido_origen.verificado_bodega and cantidad_a_despachar_str is not None and cantidad_a_despachar_str.strip().isdigit()):
                            detalle_pedido_origen.cantidad_verificada = nuevo_total_verificado_pedido
                            detalle_pedido_origen.verificado_bodega = True 
                            detalle_pedido_origen.save(update_fields=['cantidad_verificada', 'verificado_bodega'])
                            hubo_cambios_generales_en_detalle_pedido = True
                            print(f"    DetallePedido ID {detalle_pedido_origen.id} ACTUALIZADO. Nuevo Total Verificado: {detalle_pedido_origen.cantidad_verificada}")

                    except ValueError as e_val_item:
                        error_msg_item = f"Error en ítem {detalle_pedido_origen.producto.nombre} ({detalle_pedido_origen.producto.referencia}): {e_val_item}"
                        messages.error(request, error_msg_item)
                        raise transaction.TransactionManagementError(error_msg_item)
                
                print("Fin del bucle for detalles.")

                
                comprobante_creado = None
                if items_efectivamente_despachados_para_comprobante:
                    
                    comprobante_creado = ComprobanteDespacho.objects.create(
                        empresa=empresa_actual,
                        pedido=pedido,
                        fecha_hora_despacho=timezone.now(),
                        usuario_responsable=request.user,
                    )
                    print(f"ComprobanteDespacho ID {comprobante_creado.pk} CREADO para empresa '{empresa_actual.nombre}'.")                                 
                                                            
                    if comprobante_creado:
                        from factura.models import EstadoFacturaDespacho # Importar el modelo
                        EstadoFacturaDespacho.objects.create(
                            empresa=empresa_actual,
                            despacho=comprobante_creado
                        )
                        print(f"EstadoFacturaDespacho CREADO para Comprobante ID {comprobante_creado.pk}")
                                   
                    for item_data in items_efectivamente_despachados_para_comprobante:
                        DetalleComprobanteDespacho.objects.create(
                            comprobante_despacho=comprobante_creado,
                            producto=item_data['producto'],
                            cantidad_despachada=item_data['cantidad_despachada'],
                            detalle_pedido_origen=item_data['detalle_pedido_origen']
                        )
                    print(f"  {len(items_efectivamente_despachados_para_comprobante)} Detalles de ComprobanteDespacho CREADOS.")
                    messages.success(request, f"Comprobante de Despacho #{comprobante_creado.pk} generado con {len(items_efectivamente_despachados_para_comprobante)} ítem(s).")


                elif hubo_cambios_generales_en_detalle_pedido: 
                    messages.info(request, "Se actualizaron las cantidades verificadas del pedido, pero no se generó un nuevo comprobante de despacho (0 ítems despachados esta vez).")
                else: 
                    messages.info(request, "No se despacharon ítems ni se realizaron cambios en la verificación del pedido.")

                
                if hubo_cambios_generales_en_detalle_pedido or comprobante_creado: 
                    pedido.refresh_from_db() 
                    detalles_refrescados = pedido.detalles.all()
                    todas_completas = all((d.cantidad_verificada or 0) >= d.cantidad for d in detalles_refrescados)
                    
                    nuevo_estado_pedido = pedido.estado
                    if todas_completas:
                        if pedido.estado not in ['COMPLETADO', 'ENVIADO', 'ENTREGADO', 'CANCELADO']:
                            nuevo_estado_pedido = 'COMPLETADO' 
                            messages.success(request, f'Pedido #{pedido.numero_pedido_empresa} marcado como COMPLETAMENTE DESPACHADO/VERIFICADO!')
                    else: 
                        if items_efectivamente_despachados_para_comprobante or hubo_cambios_generales_en_detalle_pedido : 
                            if pedido.estado in ['PENDIENTE', 'APROBADO_ADMIN', 'PENDIENTE_BODEGA']: # Ajustar estados según tu flujo
                                nuevo_estado_pedido = 'PROCESANDO'
                                messages.info(request, f'Pedido #{pedido.numero_pedido_empresa} ahora en estado PROCESANDO.')
                            elif pedido.estado == 'PROCESANDO':
                                messages.info(request, f'Verificación/Despacho actualizado para pedido #{pedido.numero_pedido_empresa} (sigue en PROCESANDO).')
                    
                    if nuevo_estado_pedido != pedido.estado:
                        pedido.estado = nuevo_estado_pedido
                        pedido.save(update_fields=['estado'])

                
                print("Fin bloque 'with transaction.atomic()'. Transacción exitosa. Redirigiendo...")
                
                return redirect('bodega:verificar_pedido', pk=pedido.pk)

        except transaction.TransactionManagementError as e_trans:
            print(f"TransactionManagementError: {e_trans}. Rollback realizado.")
            
        except ValueError as e_val: 
            print(f"ValueError: {e_val}")
            messages.error(request, str(e_val))
        except Exception as e_global:
            print(f"Error global inesperado (POST): {type(e_global).__name__} - {e_global}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error inesperado al procesar la verificación/despacho: {e_global}")

        
        print("POST procesado con errores o rollback. Redirigiendo a GET para mostrar errores...")
        return redirect('bodega:verificar_pedido', pk=pedido.pk)

    else:
        return HttpResponse("Método no permitido.", status=405)
    
  

@login_required
@permission_required('bodega.view_comprobantedespacho', login_url='core:acceso_denegado')
def vista_imprimir_comprobante_especifico(request, pk_comprobante): 
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
   
    try:
        
        queryset = ComprobanteDespacho.objects.select_related(
            'pedido', 
            'pedido__cliente',
            'pedido__vendedor__user', 
            'usuario_responsable' 
        ).prefetch_related(
            'detalles__producto' 
        )
        comprobante = get_object_or_404(
            queryset, 
            pk=pk_comprobante, 
            pedido__empresa=empresa_actual # ¡Este filtro es la clave de la seguridad aquí!
        )
        pedido_asociado = comprobante.pedido
        
    except ComprobanteDespacho.DoesNotExist:
        messages.error(request, f"El comprobante de despacho con ID #{pk_comprobante} no existe.")
        return redirect('factura:lista_despachos_a_facturar') 

    detalles_del_comprobante = comprobante.detalles.all()
    
    # --- INICIO: Cargar Mapeo de Tallas (Refactor) ---
    TALLAS_MAPEO = empresa_actual.talla_mapeo or {}
    # --- FIN: Cargar Mapeo de Tallas ---

    if not detalles_del_comprobante:
        messages.warning(request, f"El Comprobante de Despacho #{comprobante.pk} no tiene ítems detallados.")
        return redirect('factura:detalle_despacho_factura', pk_despacho=comprobante.pk)


    items_agrupados_dict = defaultdict(lambda: {
        'nombre_producto': '', 
        'color': '',
        'tallas_cantidades': [], 
        'total_cantidad_referencia_color': 0 
    })

    for detalle_item in detalles_del_comprobante:
        producto_obj = detalle_item.producto
        clave_agrupacion = (producto_obj.referencia, producto_obj.color)

        if not items_agrupados_dict[clave_agrupacion]['nombre_producto']:
            items_agrupados_dict[clave_agrupacion]['nombre_producto'] = producto_obj.nombre 
            items_agrupados_dict[clave_agrupacion]['color'] = producto_obj.color if producto_obj.color else "-"

 # --- Aplicar Mapeo de Talla (Refactor) ---
        talla_original = producto_obj.talla if producto_obj.talla else "-"
        talla_como_texto = str(talla_original).strip()
        talla_display = TALLAS_MAPEO.get(talla_como_texto, talla_como_texto)
        # --- Fin Mapeo ---
        
        items_agrupados_dict[clave_agrupacion]['tallas_cantidades'].append({
            'talla': talla_display, # <-- LÍNEA MODIFICADA
            'cantidad': detalle_item.cantidad_despachada
        })

        items_agrupados_dict[clave_agrupacion]['total_cantidad_referencia_color'] += detalle_item.cantidad_despachada

    items_agrupados_para_pdf = []
    for (referencia, color_agrupado), data in items_agrupados_dict.items():
        data['tallas_cantidades'].sort(key=lambda x: x.get('talla', ''))
        items_agrupados_para_pdf.append({
            'referencia': referencia,
            'nombre_producto': data['nombre_producto'],
            'color': color_agrupado if color_agrupado else "-",
            'tallas_cantidades': data['tallas_cantidades'],
            'total_cantidad_referencia_color': data['total_cantidad_referencia_color']
        })

    items_agrupados_para_pdf.sort(key=lambda x: x['referencia'])
    
    print(f"Procesados {len(items_agrupados_para_pdf)} grupos de ítems para el PDF del Comprobante #{comprobante.pk}.")

    context_pdf = {
        'comprobante_despacho': comprobante, 
        'pedido': pedido_asociado,           
        'items_despachados_agrupados': items_agrupados_para_pdf,
        'fecha_despacho': comprobante.fecha_hora_despacho, 
        'logo_base64': get_logo_base_64_despacho(empresa=empresa_actual),
        'total_a_mostrar': sum(item['total_cantidad_referencia_color'] for item in items_agrupados_para_pdf)
    }
    
    filename = f"Comprobante_Despacho_Empresa{empresa_actual.pk}_{comprobante.pk}"
    
    return render_pdf_weasyprint(
        request,
        'bodega/comprobante_despacho_pdf.html', 
        context_pdf,
        filename_prefix=filename 
    )

@login_required
@permission_required('bodega.view_comprobantedespacho', login_url='core:acceso_denegado')
def vista_generar_ultimo_comprobante_pedido(request, pk): # pk_pedido es el ID del Pedido
    """
    Genera el PDF para el ÚLTIMO ComprobanteDespacho asociado a un Pedido específico.
    PK es el ID del Pedido.
    """
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    print(f"--- VISTA_GENERAR_ULTIMO_COMPROBANTE_PEDIDO: INICIO GET para Pedido ID #{pk} ---")
    
    pedido_obj = get_object_or_404(Pedido, pk=pk, empresa=empresa_actual)

    try:
        comprobante = ComprobanteDespacho.objects.select_related(
            'pedido', 
            'pedido__cliente',
            'pedido__vendedor__user', 
            'usuario_responsable' 
        ).prefetch_related(
            'detalles__producto' 
        ).filter(pedido=pedido_obj).latest('fecha_hora_despacho')
        
        pedido_asociado = comprobante.pedido 

    except ComprobanteDespacho.DoesNotExist:
        messages.error(request, f"No se encontró ningún comprobante de despacho para el Pedido #{pedido_obj.pk}.")
        return redirect('bodega:despacho_pedido', pk=pedido_obj.pk)
    except Exception as e:
        messages.error(request, f"Ocurrió un error inesperado buscando el comprobante: {e}")
        return redirect('bodega:despacho_pedido', pk=pedido_obj.pk)

    detalles_del_comprobante = comprobante.detalles.all()

    if not detalles_del_comprobante:
        messages.warning(request, f"El Comprobante de Despacho #{comprobante.pk} (último para el Pedido #{pedido_obj.pk}) no tiene ítems detallados.")
        return redirect('bodega:despacho_pedido', pk=pedido_obj.pk)
    
    # --- INICIO: Cargar Mapeo de Tallas (Refactor) ---
    TALLAS_MAPEO = empresa_actual.talla_mapeo or {}
    # --- FIN: Cargar Mapeo de Tallas ---

    items_agrupados_dict = defaultdict(lambda: {
        'nombre_producto': '', 
        'color': '',
        'tallas_cantidades': [], 
        'total_cantidad_referencia_color': 0 
    })

    for detalle_item in detalles_del_comprobante:
        producto_obj = detalle_item.producto
        clave_agrupacion = (producto_obj.referencia, producto_obj.color)

        if not items_agrupados_dict[clave_agrupacion]['nombre_producto']:
            items_agrupados_dict[clave_agrupacion]['nombre_producto'] = producto_obj.nombre 
            items_agrupados_dict[clave_agrupacion]['color'] = producto_obj.color if producto_obj.color else "-"

# --- Aplicar Mapeo de Talla (Refactor) ---
        talla_original = producto_obj.talla if producto_obj.talla else "-"
        talla_como_texto = str(talla_original).strip()
        talla_display = TALLAS_MAPEO.get(talla_como_texto, talla_como_texto)
        # --- Fin Mapeo ---

        items_agrupados_dict[clave_agrupacion]['tallas_cantidades'].append({
            'talla': talla_display, # <-- LÍNEA MODIFICADA
            'cantidad': detalle_item.cantidad_despachada
        })
        items_agrupados_dict[clave_agrupacion]['total_cantidad_referencia_color'] += detalle_item.cantidad_despachada

    items_agrupados_para_pdf = []
    for (referencia, color_agrupado), data in items_agrupados_dict.items():
        data['tallas_cantidades'].sort(key=lambda x: str(x.get('talla', '')))
        items_agrupados_para_pdf.append({
            'referencia': referencia,
            'nombre_producto': data['nombre_producto'],
            'color': color_agrupado if color_agrupado else "-",
            'tallas_cantidades': data['tallas_cantidades'],
            'total_cantidad_referencia_color': data['total_cantidad_referencia_color']
        })

    items_agrupados_para_pdf.sort(key=lambda x: x['referencia'])
    
    print(f"Procesados {len(items_agrupados_para_pdf)} grupos de ítems para el PDF del Comprobante #{comprobante.pk} (Pedido #{pedido_obj.pk}).")

    context_pdf = {
        'comprobante_despacho': comprobante, 
        'pedido': pedido_asociado,          
        'items_despachados_agrupados': items_agrupados_para_pdf,
        'fecha_despacho': comprobante.fecha_hora_despacho, 
        'logo_base64': get_logo_base_64_despacho(empresa=empresa_actual),
        'total_a_mostrar': sum(item['total_cantidad_referencia_color'] for item in items_agrupados_para_pdf)
    }
    
    filename = f"Comprobante_Despacho_Empresa{empresa_actual.pk}_P{pedido_obj.pk}_C{comprobante.pk}"
    
    return render_pdf_weasyprint(
        request,
        'bodega/comprobante_despacho_pdf.html', 
        context_pdf,
        filename_prefix=filename
    )


@login_required
@permission_required(['bodega.add_ingresobodega', 'bodega.add_movimientoinventario'], login_url='core:acceso_denegado')
def vista_registrar_ingreso(request):
    """
    Maneja el registro de un nuevo Ingreso a Bodega con sus detalles.
    """
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    if request.method == 'POST':
        form = IngresoBodegaForm(request.POST, empresa=empresa_actual)
        formset = DetalleIngresoFormSet(request.POST, prefix='detalles_ingreso', form_kwargs={'empresa': empresa_actual})


        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    ingreso_header = form.save(commit=False)
                    ingreso_header.usuario = request.user
                    ingreso_header.fecha_hora = timezone.now()
                    ingreso_header.empresa = empresa_actual
                    ingreso_header.save()

                    formset.instance = ingreso_header
                    formset.save()

                    messages.success(request, f"Ingreso a Bodega #{ingreso_header.pk} registrado exitosamente.")
                    return redirect('bodega:vista_registrar_ingreso') # Nombre corregido

            except Exception as e:
                messages.error(request, f"Error al guardar el ingreso: {e}")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else: 
        form = IngresoBodegaForm(empresa=empresa_actual)
        formset = DetalleIngresoFormSet(prefix='detalles_ingreso', form_kwargs={'empresa': empresa_actual})

    context = {
        'form': form,
        'formset': formset,
        'titulo': f'Registrar Ingreso a Bodega ({empresa_actual.nombre})'
    }
    return render(request, 'bodega/registrar_ingreso.html', context)


#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

@login_required
@permission_required('bodega.view_cabeceraconteo', login_url='core:acceso_denegado')
def exportar_plantilla_conteo(request, file_format='xlsx'): # Añadimos file_format
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido.")
        return redirect('core:index')

    # Obtenemos solo los productos de la empresa actual
    queryset = Producto.objects.filter(empresa=empresa_actual, activo=True)
    
    # Usamos el Resource, tal como lo haces en la app de clientes
    plantilla_resource = PlantillaConteoResource()
    dataset = plantilla_resource.export(queryset)
    
    # Lógica de respuesta idéntica a la tuya
    if file_format == 'csv':
        response_content = dataset.csv
        content_type = 'text/csv'
        filename = f'plantilla_conteo_{timezone.now().strftime("%Y%m%d")}.csv'
    elif file_format == 'xls':
        response_content = dataset.xls
        content_type = 'application/vnd.ms-excel'
        filename = f'plantilla_conteo_{timezone.now().strftime("%Y%m%d")}.xls'
    else: # xlsx por defecto
        response_content = dataset.xlsx
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'plantilla_conteo_{timezone.now().strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(response_content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


logger = logging.getLogger(__name__)



@transaction.atomic
def _procesar_y_guardar_conteo(request, empresa, datos_generales, datos_conteo):
    cabecera = CabeceraConteo.objects.create(
        empresa=empresa,
        fecha_conteo=datos_generales['fecha_actualizacion_stock'],
        motivo=datos_generales['motivo_conteo'],
        revisado_con=datos_generales['revisado_con'],
        notas_generales=datos_generales['notas_generales'],
        usuario_registro=request.user
    )
    stats = {'ajustados': 0, 'sin_cambio': 0}
    bodega_principal = Bodega.objects.principal(empresa)
    for producto in Producto.objects.filter(empresa=empresa, activo=True):
        cantidad_fisica = datos_conteo.get(producto.pk)
        if cantidad_fisica is not None and cantidad_fisica >= 0:

            # Lógica original: Compara contra el Stock Disponible
            stock_sistema = producto.stock_actual
            diferencia = cantidad_fisica - stock_sistema

            ConteoInventario.objects.create(
                empresa=empresa, cabecera_conteo=cabecera, producto=producto,
                cantidad_sistema_antes=stock_sistema, # <-- Guarda el stock disponible
                cantidad_fisica_contada=cantidad_fisica,
                usuario_conteo=request.user
            )
            if diferencia != 0:
                tipo_movimiento = 'ENTRADA_AJUSTE' if diferencia > 0 else 'SALIDA_AJUSTE'
                # Prevenir duplicados: usar get_or_create con documento_referencia que incluya producto y conteo
                MovimientoInventario.objects.get_or_create(
                    empresa=empresa, producto=producto, bodega=bodega_principal,
                    tipo_movimiento=tipo_movimiento,
                    documento_referencia=f"Conteo ID {cabecera.pk} - Prod #{producto.pk}",
                    defaults={
                        'cantidad': diferencia,
                        'usuario': request.user,
                        'notas': f"Ajuste por conteo. Sistema: {stock_sistema}, Físico: {cantidad_fisica}"
                    }
                )
                stats['ajustados'] += 1
            else:
                stats['sin_cambio'] += 1
    return cabecera, stats


@login_required
@permission_required('bodega.view_cabeceraconteo', login_url='core:acceso_denegado')
def vista_conteo_inventario(request):
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "No se pudo identificar la empresa.")
        return redirect('core:index')

    user = request.user
    puede_guardar = user.has_perm('bodega.add_cabeceraconteo')

    # --- LÓGICA POST (Original de tu archivo views.py.save) ---
    if request.method == 'POST':
        if not puede_guardar:
            messages.error(request, "No tienes permiso para guardar conteos y ajustar el stock.")
            return redirect('bodega:vista_conteo_inventario')

        info_form = InfoGeneralConteoForm(request.POST)
        import_form = ImportarConteoForm(request.POST, request.FILES)

        if info_form.is_valid():
            action = request.POST.get('action')
            datos_conteo = {}

            if action == 'guardar_manual':
                logger.info(f"Intento de guardado manual por '{user.username}'.")
                for key, value in request.POST.items():
                     if key.startswith('cantidad_fisica_') and value.strip().isdigit():
                        producto_id = int(key.split('_')[-1])
                        datos_conteo[producto_id] = int(value)
            
            elif action == 'importar_excel':
                 logger.info(f"Intento de importación por archivo por '{user.username}'.")
                
                 if not import_form.is_valid():
                    messages.error(request, "El formulario de importación tiene errores.")
                    return redirect('bodega:vista_conteo_inventario')
                
                 if 'archivo_conteo' not in request.FILES:
                    messages.error(request, "Para importar, primero debes seleccionar un archivo.")
                    return redirect('bodega:vista_conteo_inventario')
                
                 archivo_importado = request.FILES['archivo_conteo']
                 nombre_archivo = archivo_importado.name.lower()

                 try:
                     if nombre_archivo.endswith('.csv'):
                        contenido_str = archivo_importado.read().decode('utf-8')
                        dialect = csv.Sniffer().sniff(contenido_str.splitlines()[0], delimiters=',;')
                        reader = csv.reader(contenido_str.splitlines(), dialect)
                        next(reader)
                        for row in reader:
                            if len(row) < 7: continue
                            try:
                                p_id, c_fisica = int(row[0]), int(row[6])
                                if p_id > 0 and c_fisica >= 0: datos_conteo[p_id] = c_fisica
                            except (ValueError, TypeError, IndexError): continue

                     elif nombre_archivo.endswith(('.xlsx', '.xls')):
                        workbook = openpyxl.load_workbook(archivo_importado, data_only=True)
                        sheet = workbook[workbook.sheetnames[0]]
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if len(row) < 7: continue
                            try:
                                p_id, c_fisica = int(row[0]), int(row[6])
                                if p_id > 0 and c_fisica >= 0: datos_conteo[p_id] = c_fisica
                            except (ValueError, TypeError, IndexError): continue
                     else:
                        messages.error(request, "Formato de archivo no soportado. Sube un .csv o .xlsx.")
                        return redirect('bodega:vista_conteo_inventario')

                 except Exception as e:
                    messages.error(request, f"Error crítico durante la lectura del archivo: {e}")
                    logger.error(f"Error de parsing para conteo: {e}", exc_info=True)
                    return redirect('bodega:vista_conteo_inventario')
            
            if not datos_conteo:
                messages.warning(request, "No se encontraron datos de cantidades para procesar, ni en la tabla ni en el archivo.")
                return redirect('bodega:vista_conteo_inventario')
            
            try:
                # Usamos la lógica de guardado original (Paso 2)
                cabera, stats = _procesar_y_guardar_conteo(request, empresa_actual, info_form.cleaned_data, datos_conteo)
                if stats['ajustados'] > 0:
                    messages.success(request, f"Conteo #{cabera.pk} guardado. Se ajustó el stock de {stats['ajustados']} producto(s).")
                if stats['sin_cambio'] > 0 and stats['ajustados'] == 0:
                    messages.info(request, f"Conteo #{cabera.pk} guardado. {stats['sin_cambio']} producto(s) no tuvieron cambios de stock.")
                return redirect('bodega:descargar_informe_conteo', cabecera_id=cabera.pk)
            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado al guardar el conteo: {e}")
                logger.critical(f"Error fatal al llamar a _procesar_y_guardar_conteo: {e}", exc_info=True)
                return redirect('bodega:vista_conteo_inventario')
        else:
            messages.error(request, "Por favor corrige los errores en la información general del conteo.")

    # --- LÓGICA GET (Esta es la original que coincide con tu plantilla) ---
    
    items_a_contar = Producto.objects.filter(empresa=empresa_actual, activo=True).order_by('referencia', 'nombre', 'color', 'talla')
    
    info_form = InfoGeneralConteoForm()
    import_form = ImportarConteoForm()

    context = {
        'items_para_conteo': items_a_contar, # <--- Pasamos los productos directamente
        'titulo': f"Conteo de Inventario Físico ({empresa_actual.nombre})",
        'info_form': info_form,
        'import_form': import_form,
        'puede_guardar': puede_guardar,
    }
    return render(request, 'bodega/conteo_inventario.html', context)
    
    
@login_required
@permission_required('bodega.view_cabeceraconteo', login_url='core:acceso_denegado')
def descargar_informe_conteo(request, cabecera_id):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    cabecera = get_object_or_404(CabeceraConteo, pk=cabecera_id, empresa=empresa_actual)
    detalles_conteo = ConteoInventario.objects.filter(cabecera_conteo=cabecera).select_related('producto')
    inconsistencias = [d for d in detalles_conteo if d.diferencia != 0]

    context = {
        'cabecera': cabecera,
        'inconsistencias': inconsistencias, # Solo los que tuvieron diferencia
        'detalles_completos': detalles_conteo, # Todos los detalles para un informe completo si se desea
        'logo_base64': get_logo_base_64_despacho(empresa=empresa_actual),
        'fecha_generacion': timezone.now(),
    }

    html_string = render_to_string('bodega/conteo_informe_pdf.html', context)
    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        pdf_file = html.write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'Informe_conteo_Empresa{empresa_actual.pk}_{cabecera.pk}_{cabecera.fecha_conteo.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        print(f"Error generando PDF con WeasyPrint para conteo {cabecera.pk}: {e}")
        messages.error(request, f"Error al generar el informe PDF: {e}")
        return redirect('bodega:vista_conteo_inventario')


class InformeDespachosView(TenantAwareMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Pedido
    template_name = 'bodega/informe_despachos.html' 
    context_object_name = 'lista_pedidos'
    paginate_by = 25
    
    permission_required = 'informes.view_comprobantes_despacho'
    login_url = 'core:acceso_denegado'

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para acceder a este informe.")
        return redirect('core:index')

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        empresa_actual = self.request.tenant

        # CORRECCIÓN: Las importaciones se mueven aquí, fuera del bloque try/except,
        # para que Pylance y otros analizadores las reconozcan siempre.
        from core.auth_utils import es_vendedor, es_admin_sistema, es_factura, es_cartera
        from vendedores.models import Vendedor

        if es_vendedor(user) and not (user.is_superuser or es_admin_sistema(user) or es_factura(user) or es_cartera(user)):
            try:
                vendedor_actual = Vendedor.objects.get(user=user, user__empresa=empresa_actual)
                queryset = queryset.filter(vendedor=vendedor_actual)
            except Vendedor.DoesNotExist:
                messages.warning(self.request, "Tu usuario no está asociado a un perfil de vendedor.")
                return Pedido.objects.none()
        
        # --- SECCIÓN DE FILTRADO POR PARÁMETROS GET (INTACTA) ---
        nit_cliente_query = self.request.GET.get('nit_cliente', '').strip()
        nombre_cliente_query = self.request.GET.get('nombre_cliente', '').strip()
        numero_pedido_query = self.request.GET.get('numero_pedido', '').strip()
        fecha_pedido_inicio_query = self.request.GET.get('fecha_pedido_inicio', '').strip()
        fecha_pedido_fin_query = self.request.GET.get('fecha_pedido_fin', '').strip()
        estado_pedido_query = self.request.GET.get('estado_pedido', '').strip()

        if nit_cliente_query:
            queryset = queryset.filter(cliente__identificacion__icontains=nit_cliente_query)
        if nombre_cliente_query:
            queryset = queryset.filter(cliente__nombre_completo__icontains=nombre_cliente_query)
        if numero_pedido_query:
            try:
                pedido_pk = int(numero_pedido_query)
                queryset = queryset.filter(pk=pedido_pk)
            except (ValueError, TypeError):
                pass 
        if fecha_pedido_inicio_query and fecha_pedido_fin_query:
            queryset = queryset.filter(fecha_hora__date__range=[fecha_pedido_inicio_query, fecha_pedido_fin_query])
        elif fecha_pedido_inicio_query:
            queryset = queryset.filter(fecha_hora__date__gte=fecha_pedido_inicio_query)
        elif fecha_pedido_fin_query:
            queryset = queryset.filter(fecha_hora__date__lte=fecha_pedido_fin_query)
        
        if estado_pedido_query:
            queryset = queryset.filter(estado=estado_pedido_query)
            
        # --- OPTIMIZACIÓN Y RETORNO FINAL (INTACTO) ---
        return queryset.select_related('cliente', 'vendedor__user').annotate(
            ultima_fecha_despacho=Max('comprobantes_despacho__fecha_hora_despacho')
        ).order_by('-fecha_hora').distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        empresa_actual = self.request.tenant
        
        from core.auth_utils import es_vendedor, es_admin_sistema, es_factura, es_cartera
        
        if es_vendedor(user) and not (user.is_superuser or es_admin_sistema(user) or es_factura(user) or es_cartera(user)):
            context['titulo_pagina'] = f"Mis Pedidos y Despachos ({empresa_actual.nombre})"
        else:
            context['titulo_pagina'] = f"Informe General de Despachos ({empresa_actual.nombre})"

        context['nit_cliente_query'] = self.request.GET.get('nit_cliente', '')
        context['nombre_cliente_query'] = self.request.GET.get('nombre_cliente', '')
        context['numero_pedido_query'] = self.request.GET.get('numero_pedido', '')
        context['fecha_pedido_inicio_query'] = self.request.GET.get('fecha_pedido_inicio', '')
        context['fecha_pedido_fin_query'] = self.request.GET.get('fecha_pedido_fin', '')
        context['estado_pedido_query'] = self.request.GET.get('estado_pedido', '')
        context['ESTADO_PEDIDO_CHOICES'] = Pedido.ESTADO_PEDIDO_CHOICES
        
        return context

@login_required
@permission_required('bodega.view_ingresobodega', login_url='core:acceso_denegado')
def vista_detalle_ingreso_bodega(request, pk):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    ingreso = get_object_or_404(
        IngresoBodega.objects.select_related('usuario'), 
        pk=pk, 
        empresa=empresa_actual
    )
    detalles_del_ingreso = ingreso.detalles.select_related('producto').all()

    context = {
        'ingreso': ingreso,
        'detalles_del_ingreso': detalles_del_ingreso,
        'titulo': f'Detalle del Ingreso #{ingreso.pk} ({empresa_actual.nombre})',
        'app_name': 'bodega' 
    }
    return render(request, 'bodega/detalle_ingreso_bodega.html', context)


@login_required
@permission_required(['bodega.add_salidainternacabecera', 'bodega.add_movimientoinventario'], login_url='core:acceso_denegado')
def registrar_salida_interna(request):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    if request.method == 'POST':
        form_cabecera = SalidaInternaCabeceraForm(request.POST, empresa=empresa_actual)
        formset_detalles = DetalleSalidaInternaFormSet(request.POST, prefix='detalles_salida', form_kwargs={'empresa': empresa_actual})

        if form_cabecera.is_valid() and formset_detalles.is_valid():
            try:
                with transaction.atomic():
                    cabecera = form_cabecera.save(commit=False)
                    cabecera.responsable_entrega = request.user
                    cabecera.empresa = empresa_actual
                    if not cabecera.bodega_origen_id:
                        cabecera.bodega_origen = Bodega.objects.principal(empresa_actual)

                    if cabecera.tipo_salida == 'DONACION_BAJA':
                        cabecera.estado = 'CERRADA'
                    else:
                        # Para otros tipos, el estado por defecto 'DESPACHADA' ya está en el modelo.
                        # Si no hay fecha_prevista_devolucion y no es DONACION_BAJA, podría ser 'CERRADA' también?
                        # Esto depende de tu lógica de negocio. Por ahora, se mantiene simple.
                        pass
                    cabecera.save()

                    formset_detalles.instance = cabecera
                    detalles_guardados = formset_detalles.save() 

                    for detalle in detalles_guardados:
                        # Asegurarse que el detalle no fue marcado para borrarse y tiene cantidad
                        if detalle and detalle.cantidad_despachada > 0:
                            tipo_mov_str = f"SALIDA_{cabecera.tipo_salida.upper()}"

                            # Mapeo explícito para asegurar consistencia con MovimientoInventario.TIPO_MOVIMIENTO_CHOICES
                            tipo_mov_map = {
                                'MUESTRARIO': 'SALIDA_MUESTRARIO',
                                'EXHIBIDOR': 'SALIDA_EXHIBIDOR',
                                'TRASLADO_INTERNO': 'SALIDA_TRASLADO',
                                'PRESTAMO': 'SALIDA_PRESTAMO',
                                'DONACION_BAJA': 'SALIDA_DONACION_BAJA',
                                'OTRO': 'SALIDA_INTERNA_OTRA', # Asegúrate que 'SALIDA_INTERNA_OTRA' exista en choices
                            }
                            tipo_mov_str = tipo_mov_map.get(cabecera.tipo_salida, 'SALIDA_INTERNA_OTRA')

                            # Usar get_or_create() con documento_referencia específico por producto
                            # para evitar violación del constraint unique_together
                            MovimientoInventario.objects.get_or_create(
                                empresa=empresa_actual,
                                producto=detalle.producto,
                                bodega=cabecera.bodega_origen,
                                tipo_movimiento=tipo_mov_str,
                                documento_referencia=f"SalidaInt #{cabecera.pk}-{detalle.producto.id}",
                                defaults={
                                    'cantidad': -detalle.cantidad_despachada,
                                    'usuario': request.user,
                                    'notas': f"Salida por {cabecera.get_tipo_salida_display()} a {cabecera.destino_descripcion}"
                                }
                            )
                    
                    messages.success(request, f"Salida Interna #{cabecera.pk} registrada exitosamente. Stock actualizado.")
                    return redirect('bodega:detalle_salida_interna', pk=cabecera.pk) 

            except Exception as e:
                messages.error(request, f"Error al guardar la salida interna: {e}")
        else:
            # Errores en form_cabecera o formset_detalles
            if not form_cabecera.is_valid():
                 messages.error(request, "Por favor corrige los errores en los datos generales de la salida.")
            if not formset_detalles.is_valid():
                 messages.error(request, "Por favor corrige los errores en los productos a despachar.")
                 # Podrías iterar sobre formset_detalles.errors para dar mensajes más específicos
                 for form_error in formset_detalles.errors:
                     print(f"Error en formset: {form_error}") # Para depuración
                 for i, form_detalle in enumerate(formset_detalles):
                    if form_detalle.errors:
                        for field, error_list in form_detalle.errors.items():
                            for error in error_list:
                                messages.warning(request, f"Error en Producto #{i+1} ({field}): {error}")


    else: # GET
        form_cabecera = SalidaInternaCabeceraForm(empresa=empresa_actual)
        formset_detalles = DetalleSalidaInternaFormSet(prefix='detalles_salida', form_kwargs={'empresa': empresa_actual})

    context = {
        'form_cabecera': form_cabecera,
        'formset_detalles': formset_detalles,
        'titulo': f'Registrar Nueva Salida Interna ({empresa_actual.nombre})'
    }
    return render(request, 'bodega/registrar_salida_interna.html', context)

@login_required
@permission_required('bodega.view_salidainternacabecera', login_url='core:acceso_denegado')
def detalle_salida_interna(request, pk): 
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    salida_interna = get_object_or_404(
        SalidaInternaCabecera.objects.select_related('responsable_entrega'), 
        pk=pk, 
        empresa=empresa_actual
    )
    detalles_items = salida_interna.detalles.select_related('producto').all()

    context = {
        'salida_interna': salida_interna,
        'detalles_items': detalles_items,
        'titulo': f"Detalle Salida Interna #{salida_interna.pk} ({empresa_actual.nombre})"
    }
    return render(request, 'bodega/detalle_salida_interna.html', context)


@require_POST
@login_required
@permission_required('bodega.change_salidainternacabecera', login_url='core:acceso_denegado')
def cerrar_salida_interna(request, pk):
    """
    Cambia el estado de una Salida Interna a 'CERRADA'.
    Esto significa que los productos no serán devueltos y la operación se da por finalizada.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido.")
        return redirect('core:index')

    salida = get_object_or_404(SalidaInternaCabecera, pk=pk, empresa=empresa_actual)
    
    ESTADOS_PERMITIDOS_PARA_CERRAR = ['DESPACHADA', 'DEVUELTA_PARCIAL']

    if salida.estado in ESTADOS_PERMITIDOS_PARA_CERRAR:        
        salida.estado = 'CERRADA'
        salida.save(update_fields=['estado'])
        messages.success(request, f"La Salida Interna #{salida.pk} ha sido cerrada exitosamente. El stock ya fue descontado previamente.")
    else:
        messages.warning(request, f"La Salida Interna #{salida.pk} no se puede cerrar porque su estado actual es '{salida.get_estado_display()}'.")
    
    return redirect('bodega:detalle_salida_interna', pk=salida.pk)


@login_required
@permission_required('bodega.view_salidainternacabecera', login_url='core:acceso_denegado')
def generar_pdf_salida_interna(request, pk):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    salida_interna = get_object_or_404(
        SalidaInternaCabecera.objects.select_related('responsable_entrega'), 
        pk=pk, 
        empresa=empresa_actual
    )
    detalles_items = salida_interna.detalles.select_related('producto').all()
    
    logo_para_pdf = get_logo_base_64_despacho(empresa=empresa_actual)

    context_pdf = {
        'salida_interna': salida_interna,
        'detalles_items': detalles_items,
        'logo_base64': logo_para_pdf,
        'fecha_generacion': timezone.now(),
    }
    
    filename = f"Comprobante_Despacho_Empresa{empresa_actual.pk}_{salida_interna.pk}"
    
    return render_pdf_weasyprint(
        request,
        'bodega/salida_interna_pdf.html',
        context_pdf,
        filename_prefix=filename
    )

@login_required
@permission_required('bodega.view_salidainternacabecera', login_url='core:acceso_denegado')
def lista_salidas_internas(request):
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    salidas_query = SalidaInternaCabecera.objects.filter(
        empresa=empresa_actual
    ).select_related('responsable_entrega')

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_salida_filtro = request.GET.get('tipo_salida_filtro')
    estado_filtro = request.GET.get('estado_filtro') # Nuevo filtro por estado

    if fecha_inicio:
        salidas_query = salidas_query.filter(fecha_hora_salida__date__gte=fecha_inicio)
    if fecha_fin:
        salidas_query = salidas_query.filter(fecha_hora_salida__date__lte=fecha_fin)
    if tipo_salida_filtro:
        salidas_query = salidas_query.filter(tipo_salida=tipo_salida_filtro)
    if estado_filtro:
        salidas_query = salidas_query.filter(estado=estado_filtro)


    salidas = salidas_query.order_by('-fecha_hora_salida')
    
    context = {
        'salidas_list': salidas,
        'titulo': f"Salidas Internas de Bodega ({empresa_actual.nombre})",
        'TIPO_SALIDA_CHOICES': SalidaInternaCabecera.TIPO_SALIDA_CHOICES, # Para el dropdown de filtro
        'ESTADO_SALIDA_CHOICES': SalidaInternaCabecera.ESTADO_SALIDA_CHOICES, # Para el dropdown de filtro de estado
        'request_get': request.GET # Para mantener los valores de los filtros en la plantilla
    }
    return render(request, 'bodega/lista_salidas_internas.html', context)

@login_required
@permission_required(['bodega.change_salidainternacabecera', 'bodega.add_movimientoinventario'], login_url='core:acceso_denegado')
def registrar_devolucion_salida_interna(request, pk_cabecera):     
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    salida_interna = get_object_or_404(
        SalidaInternaCabecera.objects.prefetch_related('detalles__producto'), 
        pk=pk_cabecera,
        empresa=empresa_actual
    )

    if salida_interna.estado in ['CERRADA', 'DEVUELTA_TOTALMENTE']:
        messages.warning(request, f"La Salida Interna #{salida_interna.pk} ya está {salida_interna.get_estado_display()} y no admite más devoluciones.")
        return redirect('bodega:detalle_salida_interna', pk=salida_interna.pk)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                algo_devuelto_en_esta_transaccion = False
                total_items_con_devolucion_valida_esta_vez = 0

                for detalle_salida in salida_interna.detalles.all():
                    input_name = f'cantidad_devuelta_{detalle_salida.pk}'
                    cantidad_a_devolver_str = request.POST.get(input_name, '0')
                    
                    try:
                        cantidad_a_devolver_ahora = int(cantidad_a_devolver_str)
                    except ValueError:
                        messages.error(request, f"Valor inválido para cantidad a devolver del producto '{detalle_salida.producto}'. Use números enteros.")
                        raise transaction.TransactionManagementError("Valor de cantidad inválido.")

                    if cantidad_a_devolver_ahora < 0:
                        messages.error(request, f"La cantidad a devolver para '{detalle_salida.producto}' no puede ser negativa.")
                        raise transaction.TransactionManagementError("Cantidad negativa.")
                    
                    pendiente_devolucion_item = detalle_salida.cantidad_pendiente_devolucion
                    if cantidad_a_devolver_ahora > pendiente_devolucion_item:
                        messages.error(request, f"Para '{detalle_salida.producto}', intenta devolver {cantidad_a_devolver_ahora} pero solo hay {pendiente_devolucion_item} pendiente(s).")
                        raise transaction.TransactionManagementError("Devolución excede cantidad pendiente.")

                    if cantidad_a_devolver_ahora > 0:
                        detalle_salida.cantidad_devuelta = F('cantidad_devuelta') + cantidad_a_devolver_ahora
                        detalle_salida.save()
                        # No es necesario refresh_from_db aquí si solo usamos F() y no leemos el valor inmediatamente en la misma transacción.
                        
                        # Mapeo para tipo de movimiento de devolución
                        tipo_mov_devolucion_map = {
                            'MUESTRARIO': 'ENTRADA_DEV_MUESTRARIO',
                            'EXHIBIDOR': 'ENTRADA_DEV_EXHIBIDOR',
                            'TRASLADO_INTERNO': 'ENTRADA_DEV_TRASLADO',
                            'PRESTAMO': 'ENTRADA_DEV_PRESTAMO',
                            'OTRO': 'ENTRADA_DEV_INTERNA_OTRA',
                        }
                        # Usar el tipo de salida original para determinar el tipo de movimiento de devolución
                        tipo_mov_str = tipo_mov_devolucion_map.get(salida_interna.tipo_salida, 'ENTRADA_DEV_INTERNA_OTRA')


                        # Prevenir duplicados: usar get_or_create con documento_referencia que incluya producto y salida
                        MovimientoInventario.objects.get_or_create(
                            empresa=empresa_actual,
                            producto=detalle_salida.producto,
                            bodega=salida_interna.bodega_origen,
                            tipo_movimiento=tipo_mov_str,
                            documento_referencia=f"Dev SalidaInt #{salida_interna.pk} - Prod #{detalle_salida.producto.pk}",
                            defaults={
                                'cantidad': cantidad_a_devolver_ahora,
                                'usuario': request.user,
                                'notas': f"Devolución de {detalle_salida.producto} de Salida Interna #{salida_interna.pk}"
                            }
                        )
                        algo_devuelto_en_esta_transaccion = True
                        total_items_con_devolucion_valida_esta_vez +=1
                
                if algo_devuelto_en_esta_transaccion:
                    salida_interna.refresh_from_db() 
                    todos_los_items_devueltos_completamente = all(
                        d.cantidad_pendiente_devolucion == 0 for d in salida_interna.detalles.all()
                    )
                    
                    if todos_los_items_devueltos_completamente:
                        salida_interna.estado = 'DEVUELTA_TOTALMENTE'
                    else:
                        salida_interna.estado = 'DEVUELTA_PARCIAL' # Se mantiene o cambia a parcial
                    salida_interna.save(update_fields=['estado'])
                    messages.success(request, f"Devolución de {total_items_con_devolucion_valida_esta_vez} ítem(s) para Salida Interna #{salida_interna.pk} registrada. Stock actualizado.")
                else:
                    messages.info(request, "No se ingresaron cantidades para devolver en ningún ítem.")

                return redirect('bodega:detalle_salida_interna', pk=salida_interna.pk)

        except transaction.TransactionManagementError:
            pass 
        except Exception as e:
            messages.error(request, f"Error inesperado al procesar la devolución: {e}")
            
    detalles_con_pendientes = []
    hay_algo_pendiente_general = False
    for d in salida_interna.detalles.all(): # Iterar sobre los detalles ya cargados
        pendiente = d.cantidad_pendiente_devolucion
        if pendiente > 0:
            hay_algo_pendiente_general = True
        detalles_con_pendientes.append({
            'detalle_id': d.pk,
            'producto_nombre': str(d.producto), # Usar el __str__ del producto
            'producto_referencia': d.producto.referencia,
            'cantidad_despachada': d.cantidad_despachada,
            'cantidad_ya_devuelta': d.cantidad_devuelta,
            'cantidad_pendiente': pendiente
        })

    context = {
        'salida_interna': salida_interna,
        'detalles_items': detalles_con_pendientes,
        'hay_algo_pendiente_general': hay_algo_pendiente_general,
        'titulo': f'Devolución de Salida Interna #{salida_interna.pk} ({empresa_actual.nombre})'
    }
    return render(request, 'bodega/registrar_devolucion_salida_interna.html', context)


@login_required
@permission_required('bodega.view_salidainternacabecera', login_url='core:acceso_denegado')
def generar_pdf_devolucion_salida_interna(request, pk_cabecera):
    """
    Genera un PDF para el comprobante de devolución de una Salida Interna.
    Muestra los ítems que han sido devueltos.
    """
    
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')
    
    salida_interna = get_object_or_404(
        SalidaInternaCabecera.objects.select_related('responsable_entrega').prefetch_related('detalles__producto'),
        pk=pk_cabecera,
        empresa=empresa_actual
    )

    detalles_para_pdf = []
    for detalle in salida_interna.detalles.all():
        detalles_para_pdf.append({
            'producto_referencia': detalle.producto.referencia,
            'producto_nombre': str(detalle.producto),
            'producto_color': detalle.producto.color or "N/A",
            'producto_talla': detalle.producto.talla or "N/A",
            'cantidad_despachada': detalle.cantidad_despachada,
            'cantidad_devuelta': detalle.cantidad_devuelta,
            'cantidad_pendiente': detalle.cantidad_pendiente_devolucion,
            'observaciones_detalle': detalle.observaciones_detalle
        })
    
    
    logo_para_pdf = get_logo_base_64_despacho(empresa=empresa_actual)

    context_pdf = {
        'salida_interna': salida_interna,
        'detalles_items_devolucion': detalles_para_pdf,
        'logo_base64': logo_para_pdf,
        'fecha_generacion': timezone.now(),
        'titulo_comprobante': f"Comprobante de Devolución - Salida Interna N° {salida_interna.pk}"
    }
    
    filename = f"Comprobante_Despacho_Empresa{empresa_actual.pk}_{salida_interna.pk}"
    
    return render_pdf_weasyprint(
        request,
        'bodega/devolucion_salida_interna_pdf.html',
        context_pdf,
        filename_prefix=filename
    )


@csrf_exempt
@require_POST
def validar_item_despacho_ajax(request):
    """
    Valida un producto escaneado contra un pedido, actualizando la cantidad despachada.
    Devuelve una respuesta JSON para ser manejada por el frontend.
    """
    try:
        data = json.loads(request.body)
        pedido_id = data.get('pedido_id')
        codigo_barras = data.get('codigo_barras')

        if not pedido_id or not codigo_barras:
            return JsonResponse({'status': 'error', 'message': 'Faltan datos (pedido o código de barras).'}, status=400)

        # 1. Buscar el producto por su código de barras
        try:
            # Asegúrate que el campo en tu modelo Producto se llame 'codigo_barras'
            producto = Producto.objects.get(codigo_barras=codigo_barras, empresa=request.tenant)
        except Producto.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': f'Código de barras "{codigo_barras}" no encontrado.'}, status=404)

        # 2. Validar si el producto pertenece al pedido
        detalle = DetallePedido.objects.filter(pedido_id=pedido_id, producto=producto).first()

        if not detalle:
            return JsonResponse({'status': 'error', 'message': f'El producto "{producto}" no pertenece a este pedido.'}, status=400)

        # 3. Verificar si aún se pueden despachar unidades de este ítem
        # Usamos el campo 'cantidad_verificada' que tienes en tu input
        if detalle.cantidad_verificada >= detalle.cantidad:
            return JsonResponse({
                'status': 'warning', # Usamos 'warning' para un estado diferente
                'message': f'Todas las unidades de "{producto}" ya fueron despachadas.',
                'detalle_id': detalle.pk
            }, status=200) # Devolvemos 200 para que el success de AJAX lo maneje

        # ¡Éxito! La validación es correcta.
        # Opcional: Aquí podrías incrementar 'cantidad_verificada' en la BD si quieres persistencia en cada escaneo.
        # detalle.cantidad_verificada += 1
        # detalle.save()

        return JsonResponse({
            'status': 'success',
            'message': f'Correcto: {producto}',
            'detalle_id': detalle.pk,
            'producto_nombre': str(producto)
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error interno del servidor: {str(e)}'}, status=500)
    
@login_required
@permission_required('bodega.view_cabeceraconteo', login_url='core:acceso_denegado')
def lista_informes_conteo(request):
    """
    Muestra una lista paginada de todos los conteos de inventario realizados.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')

    conteos_list = CabeceraConteo.objects.filter(
        empresa=empresa_actual
    ).order_by('-fecha_hora_registro')

    context = {
        'conteos_list': conteos_list,
        'titulo': f'Historial de Conteos de Inventario ({empresa_actual.nombre})',
    }
    return render(request, 'bodega/lista_informes_conteo.html', context)    

@require_POST
@login_required
@permission_required('pedidos.change_pedido', login_url='core:acceso_denegado')
def finalizar_pedido_incompleto(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')

    try:
        with transaction.atomic():
            pedido = get_object_or_404(Pedido.objects.select_for_update().prefetch_related('detalles__producto'), pk=pk, empresa=empresa_actual)

            if pedido.estado == 'ENVIADO' or pedido.estado == 'ENTREGADO' or pedido.estado == 'CANCELADO':
                messages.warning(request, f"El pedido #{pedido.numero_pedido_empresa} ya está en un estado final y no puede ser marcado como incompleto.")
                return redirect('bodega:lista_pedidos_bodega')
            
            # Eliminar cualquier borrador de despacho pendiente para este pedido
            BorradorDespacho.objects.filter(pedido=pedido, empresa=empresa_actual).delete()
            logger.info(f"DEBUG: Borrador de despacho para Pedido {pedido.pk} eliminado al finalizar incompleto.")

            # Actualizar el estado de los detalles del pedido y REINTRODUCIR MovimientoInventario para pendientes
            for detalle in pedido.detalles.all():
                cantidad_pendiente = detalle.cantidad - (detalle.cantidad_verificada or 0)
                if cantidad_pendiente > 0:
                    # REINTRODUCIR: Crear MovimientoInventario para las cantidades que VUELVEN
                    MovimientoInventario.objects.create(
                        empresa=empresa_actual,
                        producto=detalle.producto,
                        bodega=Bodega.objects.principal(empresa_actual),
                        cantidad=cantidad_pendiente, # Cantidad positiva para entrada
                        tipo_movimiento='ENTRADA_CANCELACION', # Tipo que indica que regresa por cancelación
                        documento_referencia=f"Devolución por Pedido Incompleto {pedido.pk}",
                        usuario=request.user,
                        notas=f"Devolución de {cantidad_pendiente} unidades no despachadas del pedido {pedido.pk}"
                    )
                    logger.info(f"DEBUG: MovimientoInventario de ENTRADA para {cantidad_pendiente} de {detalle.producto.referencia} (Pedido {pedido.pk}) al finalizar incompleto.")
                
                # Opcional: Podrías querer setear cantidad_verificada a su cantidad total para asegurar la "finalización" lógica
                # pero para el borrador y la lógica actual, ya estará en el valor despachado.
                # Si el objetivo es que la cantidad verificada refleje solo lo *despachado*, no la toques aquí.

            pedido.estado = 'ENVIADO_INCOMPLETO' # O un estado similar que indique el cierre
            pedido.save(update_fields=['estado'])
            messages.success(request, f"Pedido #{pedido.numero_pedido_empresa} marcado como INCOMPLETO. Las unidades pendientes han regresado al inventario.")

    except Exception as e:
        messages.error(request, f"Error al finalizar pedido incompleto: {e}")
        logger.error(f"Error en finalizar_pedido_incompleto: {e}", exc_info=True)
    
    return redirect('bodega:lista_pedidos_bodega')

@require_POST
@login_required
@permission_required('pedidos.change_pedido', login_url='core:acceso_denegado')
def cancelar_pedido_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "Acceso no válido. No se pudo identificar la empresa.")
        return redirect('core:index')

    try:
        with transaction.atomic():
            pedido = get_object_or_404(Pedido.objects.select_for_update().prefetch_related('detalles__producto'), pk=pk, empresa=empresa_actual)

            if pedido.estado == 'ENVIADO' or pedido.estado == 'ENTREGADO' or pedido.estado == 'CANCELADO':
                messages.warning(request, f"El pedido #{pedido.numero_pedido_empresa} ya está en un estado final y no puede ser cancelado.")
                return redirect('bodega:lista_pedidos_bodega')

            # Eliminar cualquier borrador de despacho pendiente para este pedido
            BorradorDespacho.objects.filter(pedido=pedido, empresa=empresa_actual).delete()
            logger.info(f"DEBUG: Borrador de despacho para Pedido {pedido.pk} eliminado al cancelar.")

            # Reintroducir MovimientoInventario para TODAS las unidades pedidas
            for detalle in pedido.detalles.all():
                cantidad_a_devolver = detalle.cantidad # Se devuelve toda la cantidad pedida
                if cantidad_a_devolver > 0:
                    # REINTRODUCIR: Crear MovimientoInventario para todas las cantidades que VUELVEN
                    MovimientoInventario.objects.create(
                        empresa=empresa_actual,
                        producto=detalle.producto,
                        bodega=Bodega.objects.principal(empresa_actual),
                        cantidad=cantidad_a_devolver, # Cantidad positiva para entrada
                        tipo_movimiento='ENTRADA_CANCELACION', # Tipo que indica que regresa por cancelación
                        documento_referencia=f"Devolución por Cancelación Pedido {pedido.pk}",
                        usuario=request.user,
                        notas=f"Devolución de {cantidad_a_devolver} unidades del pedido cancelado {pedido.pk}"
                    )
                    logger.info(f"DEBUG: MovimientoInventario de ENTRADA para {cantidad_a_devolver} de {detalle.producto.referencia} (Pedido {pedido.pk}) al cancelar.")
                
                # Opcional: Resetear cantidad_verificada a 0 al cancelar el pedido completamente
                detalle.cantidad_verificada = 0
                detalle.save(update_fields=['cantidad_verificada'])


            pedido.estado = 'CANCELADO'
            pedido.save(update_fields=['estado'])
            messages.success(request, f"Pedido #{pedido.numero_pedido_empresa} CANCELADO exitosamente. Todas las unidades han regresado al inventario.")

    except Exception as e:
        messages.error(request, f"Error al cancelar el pedido: {e}")
        logger.error(f"Error en cancelar_pedido_bodega: {e}", exc_info=True)
    
    return redirect('bodega:lista_pedidos_bodega')

@login_required
# Asegúrate que el permiso sea el adecuado para tu app, ej: 'productos.view_producto'
@permission_required('productos.view_inventory_report', login_url='core:acceso_denegado')
def vista_informe_inventario(request):
    """
    Muestra un informe completo y paginado del inventario físico en bodega,
    con opciones de filtrado.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        # Aquí le decimos al usuario exactamente qué pasa en lugar de un error 500
        messages.error(request, "No se ha podido identificar la empresa para este informe.")
        return redirect('core:index') # O donde quieras redirigir
    
    # Query base: solo productos activos de la empresa actual
    productos_queryset = Producto.objects.filter(
        empresa=empresa_actual,
        activo=True
    )

    # Lógica de filtrado
    referencia_q = request.GET.get('referencia', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    bodega_q = request.GET.get('bodega', '').strip()

    if referencia_q:
        productos_queryset = productos_queryset.filter(referencia__icontains=referencia_q)
    if nombre_q:
        productos_queryset = productos_queryset.filter(nombre__icontains=nombre_q)
    if bodega_q:
        productos_queryset = productos_queryset.filter(ubicacion_id=bodega_q)

    # Calcular totales (después de filtrar) con una sola consulta agregada,
    # en vez de recorrer 'stock_actual' producto por producto (N+1 queries).
    total_productos = productos_queryset.count()
    total_unidades = productos_queryset.aggregate(
        total=Coalesce(Sum('movimientos__cantidad'), Value(0))
    )['total']

    # Traemos todo (ya es UNA sola consulta anotada, sin N+1) y lo organizamos
    # en memoria por Género > Referencia+Color, con las tallas como columnas.
    # Esto evita mezclar sistemas de tallas incompatibles (ej. Caballero 28-38
    # con Dama 3-16) en una sola tabla plana con columnas vacías.
    productos_lista = list(
        productos_queryset.select_related('ubicacion').annotate(
            stock_actual_calculado=Coalesce(Sum('movimientos__cantidad'), Value(0))
        ).order_by('referencia', 'color', 'talla')
    )

    def _normalizar_talla(talla_val):
        if talla_val is None or str(talla_val).strip() == '':
            return '-'
        return str(talla_val).strip()

    def _talla_orden(talla_str):
        try:
            return (0, float(talla_str.replace(',', '.')))
        except ValueError:
            return (1, talla_str)

    def _normalizar_genero(genero_val):
        # Normaliza may/min y espacios (ej. "Dama"/"DAMA"/"dama" -> "DAMA").
        # No corrige errores tipográficos genuinos (ej. "DAM"/"DAAM"), esos
        # quedan en su propia sección para que se detecten y corrijan en Productos.
        return (genero_val or '').strip().upper()

    # Orden estable por género normalizado (conserva el orden referencia/color/talla
    # ya aplicado por la consulta dentro de cada grupo de género).
    productos_lista.sort(key=lambda p: _normalizar_genero(p.genero))

    GENERO_LABELS = dict(Producto.GeneroOpciones.choices)
    secciones = []
    total_referencias = 0
    generos_no_reconocidos = set()

    for genero, productos_genero_iter in groupby(productos_lista, key=lambda p: _normalizar_genero(p.genero)):
        productos_genero = list(productos_genero_iter)

        columnas_talla = sorted(
            {_normalizar_talla(p.talla) for p in productos_genero},
            key=_talla_orden
        )
        indice_talla = {t: i for i, t in enumerate(columnas_talla)}

        grupos = []
        for (ref, color), items_iter in groupby(productos_genero, key=lambda p: (p.referencia, p.color)):
            items = list(items_iter)
            cantidades = [0] * len(columnas_talla)
            ubicaciones = []
            valor_costo_grupo = Decimal('0.00')
            valor_venta_grupo = Decimal('0.00')
            ids_vistos = set()
            for p in items:
                idx = indice_talla[_normalizar_talla(p.talla)]
                cantidades[idx] += p.stock_actual_calculado
                if p.ubicacion_id and p.ubicacion_id not in ids_vistos:
                    ids_vistos.add(p.ubicacion_id)
                    ubicaciones.append(str(p.ubicacion))
                valor_costo_grupo += Decimal(p.stock_actual_calculado) * (p.costo or Decimal('0.00'))
                valor_venta_grupo += Decimal(p.stock_actual_calculado) * (p.precio_venta or Decimal('0.00'))
            grupos.append({
                'referencia': ref,
                'color': color or 'Sin Color',
                'nombre': items[0].nombre,
                'ubicacion': ', '.join(ubicaciones) if ubicaciones else '-',
                'tallas_cantidades': cantidades,
                'total': sum(cantidades),
                'valor_costo': valor_costo_grupo,
                'valor_venta': valor_venta_grupo,
            })

        total_referencias += len(grupos)
        if genero not in GENERO_LABELS:
            generos_no_reconocidos.add(genero)
        secciones.append({
            'genero': genero,
            'genero_label': GENERO_LABELS.get(genero, genero),
            'columnas_talla': columnas_talla,
            'grupos': grupos,
            'total_unidades_seccion': sum(g['total'] for g in grupos),
            'valor_costo_seccion': sum((g['valor_costo'] for g in grupos), Decimal('0.00')),
            'valor_venta_seccion': sum((g['valor_venta'] for g in grupos), Decimal('0.00')),
        })

    total_valor_costo = sum((s['valor_costo_seccion'] for s in secciones), Decimal('0.00'))
    total_valor_venta = sum((s['valor_venta_seccion'] for s in secciones), Decimal('0.00'))

    context = {
        'secciones': secciones,
        'titulo': f'Informe de Inventario Físico ({empresa_actual.nombre})',
        'total_productos': total_productos,
        'total_unidades': total_unidades,
        'generos_no_reconocidos': sorted(generos_no_reconocidos),
        'total_referencias': total_referencias,
        'total_valor_costo': total_valor_costo,
        'total_valor_venta': total_valor_venta,
        'es_administracion': es_administracion(request.user),
        'bodegas_disponibles': Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre'),
        'bodega_seleccionada': bodega_q,
        'filtros_activos': request.GET # Para mantener los filtros en la exportación
    }
    return render(request, 'bodega/informe_inventario.html', context)


@login_required
@permission_required('productos.view_producto', login_url='core:acceso_denegado')
def exportar_inventario_excel(request):
    """
    Genera y descarga un archivo Excel con el inventario actual,
    aplicando los mismos filtros que la vista del informe.
    """
    empresa_actual = getattr(request, 'tenant', None)
    productos_queryset = Producto.objects.filter(
        empresa=empresa_actual, 
        activo=True
    ).order_by('referencia', 'color', 'talla')

    # Re-aplicar los mismos filtros que en la vista principal
    referencia_q = request.GET.get('referencia', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    bodega_q = request.GET.get('bodega', '').strip()

    if referencia_q:
        productos_queryset = productos_queryset.filter(referencia__icontains=referencia_q)
    if nombre_q:
        productos_queryset = productos_queryset.filter(nombre__icontains=nombre_q)
    if bodega_q:
        productos_queryset = productos_queryset.filter(ubicacion_id=bodega_q)

    # Anotamos el stock en una sola consulta (evita N+1 al recorrer 'stock_actual' por producto)
    productos_queryset = productos_queryset.select_related('ubicacion').annotate(
        stock_actual_calculado=Coalesce(Sum('movimientos__cantidad'), Value(0))
    )

    # Crear el archivo Excel en memoria
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Inventario'

    # Escribir la cabecera
    headers = ['Referencia', 'Descripción', 'Detalle', 'Color', 'Talla', 'Ubicación', 'Cantidad en Stock']
    sheet.append(headers)

    # Escribir los datos de los productos
    for producto in productos_queryset:
        sheet.append([
            producto.referencia,
            producto.nombre,
            producto.descripcion or '',
            producto.color or '',
            producto.talla or '',
            str(producto.ubicacion) if producto.ubicacion_id else '',
            producto.stock_actual_calculado
        ])

    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="informe_inventario.xlsx"'
    workbook.save(response)

    return response


@login_required
@permission_required('productos.view_inventory_report', login_url='core:acceso_denegado')
def informe_inventario_comparativo(request):
    """
    Tabla comparativa: muestra, para cada variante de producto, cuánto stock
    hay en CADA bodega de la empresa, una al lado de la otra (columnas),
    para poder comparar de un vistazo dónde está el inventario.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "No se ha podido identificar la empresa para este informe.")
        return redirect('core:index')

    referencia_q = request.GET.get('referencia', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()

    productos_queryset = Producto.objects.filter(empresa=empresa_actual, activo=True)
    if referencia_q:
        productos_queryset = productos_queryset.filter(referencia__icontains=referencia_q)
    if nombre_q:
        productos_queryset = productos_queryset.filter(nombre__icontains=nombre_q)
    productos_queryset = productos_queryset.order_by('referencia', 'color', 'talla')

    bodegas = list(Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre'))

    paginator = Paginator(productos_queryset, 100)
    page_obj = paginator.get_page(request.GET.get('page'))
    productos_pagina = list(page_obj.object_list)

    stock_map = {}
    if productos_pagina:
        producto_ids = [p.pk for p in productos_pagina]
        filas_stock = MovimientoInventario.objects.filter(
            producto_id__in=producto_ids
        ).values('producto_id', 'bodega_id').annotate(total=Sum('cantidad'))
        for fila in filas_stock:
            stock_map[(fila['producto_id'], fila['bodega_id'])] = fila['total'] or 0

    filas = []
    for producto in productos_pagina:
        valores_por_bodega = [stock_map.get((producto.pk, b.pk), 0) for b in bodegas]
        filas.append({
            'producto': producto,
            'valores': valores_por_bodega,
            'total': sum(valores_por_bodega),
        })

    context = {
        'titulo': f'Comparativo de Stock por Bodega ({empresa_actual.nombre})',
        'bodegas': bodegas,
        'filas': filas,
        'page_obj': page_obj,
        'filtros_activos': request.GET,
    }
    return render(request, 'bodega/informe_inventario_comparativo.html', context)


class InformeMovimientoInventarioView(TenantAwareMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = MovimientoInventario
    template_name = 'bodega/informe_movimiento_inventario.html'
    context_object_name = 'movimientos_list'
    permission_required = 'bodega.view_movimientoinventario'
    login_url = 'core:acceso_denegado'
    paginate_by = 50 # Aumentamos un poco la paginación

    def get_queryset(self):
        """
        Filtra los movimientos según los parámetros GET del formulario.
        """
        empresa_actual = self.request.tenant
        queryset = super().get_queryset().filter(empresa=empresa_actual).select_related('producto', 'usuario', 'bodega')

        # --- APLICAMOS LOS NUEVOS FILTROS ---
        producto_id = self.request.GET.get('producto_id')
        bodega_id = self.request.GET.get('bodega_id')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        usuario_id = self.request.GET.get('usuario_id')

        if not producto_id and not bodega_id:
            # Si no se ha seleccionado ni producto ni bodega, no mostramos ningún movimiento.
            return queryset.none()

        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)
        if bodega_id:
            queryset = queryset.filter(bodega_id=bodega_id)

        # Filtros opcionales
        if fecha_inicio:
            queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)

        return queryset.order_by('-fecha_hora')

    def get_context_data(self, **kwargs):
        """
        Añade datos extra al contexto para el formulario de filtros y los resultados.
        """
        context = super().get_context_data(**kwargs)
        empresa_actual = self.request.tenant
        producto_id_seleccionado = self.request.GET.get('producto_id')
        bodega_id_seleccionada = self.request.GET.get('bodega_id')

        # Pasamos la lista de productos, bodegas y usuarios al template para los selectores del formulario
        context['productos_list'] = Producto.objects.filter(empresa=empresa_actual, activo=True)
        context['bodegas_list'] = Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre')
        context['usuarios_list'] = get_user_model().objects.filter(empresa=empresa_actual, is_active=True).order_by('username')

        context['titulo'] = 'Informe de Movimientos de Inventario'
        context['filtros_activos'] = self.request.GET # Para mantener los valores en el form

        titulo_partes = []
        if producto_id_seleccionado:
            producto_seleccionado = Producto.objects.filter(pk=producto_id_seleccionado).first()
            context['producto_seleccionado'] = producto_seleccionado
            if producto_seleccionado:
                titulo_partes.append(producto_seleccionado.referencia)
                # Calcular stock actual para el producto seleccionado
                stock_calculado = producto_seleccionado.movimientos.aggregate(total=Sum('cantidad'))['total']
                context['stock_actual_calculado'] = stock_calculado or 0

        if bodega_id_seleccionada:
            bodega_seleccionada = Bodega.objects.filter(pk=bodega_id_seleccionada, empresa=empresa_actual).first()
            context['bodega_seleccionada'] = bodega_seleccionada
            if bodega_seleccionada:
                titulo_partes.append(bodega_seleccionada.nombre)

        if titulo_partes:
            context['titulo'] = f'Movimientos de: {" — ".join(titulo_partes)}'

        return context

class IngresoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = IngresoBodega
    template_name = 'bodega/modificar_ingreso_bodega.html'  
    permission_required = 'bodega.change_ingresobodega'
    form_class = IngresoBodegaForm
    
    def get_queryset(self):
        """Sobrescribe el queryset para filtrar por la empresa del usuario."""
        queryset = super().get_queryset()
        empresa_actual = getattr(self.request, 'tenant', None)
        if not empresa_actual:
            logger.error("Intento de acceso a IngresoUpdateView sin empresa asignada.")
            return queryset.none()
        return queryset.filter(empresa=empresa_actual)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = DetalleIngresoModificarFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = DetalleIngresoModificarFormSet(instance=self.object)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        formset = DetalleIngresoModificarFormSet(request.POST, instance=self.object)

        if formset.is_valid():
            with transaction.atomic():
                # 1. Obtener los detalles de ingreso originales
                detalles_originales = {
                    detalle.pk: detalle.cantidad
                    for detalle in self.object.detalles.all()
                }

                # 2. Validar que NO se cambien cantidades (evita inflado de inventario)
                hay_cambios_cantidad = False
                for form in formset:
                    if form.instance.pk:
                        nuevo_detalle = form.instance
                        cantidad_antigua = detalles_originales.get(nuevo_detalle.pk, 0)

                        if nuevo_detalle.cantidad != cantidad_antigua:
                            # Revertir a la cantidad original
                            form.instance.cantidad = cantidad_antigua
                            hay_cambios_cantidad = True

                if hay_cambios_cantidad:
                    messages.warning(request, "⚠️ No se pueden modificar las cantidades de un ingreso ya registrado. "
                                   "Si necesitas cambiar cantidades, cancela el ingreso y crea uno nuevo.")

                # 3. Guardar los nuevos valores (pero las cantidades son las originales)
                formset.save()

            messages.success(request, "Los detalles del ingreso han sido actualizados exitosamente.")
            return redirect('bodega:detalle_ingreso', pk=self.object.pk)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
            return self.render_to_response(self.get_context_data(formset=formset))
        
@login_required
@permission_required('bodega.can_manage_product_changes', login_url='core:acceso_denegado')
def realizar_cambio_producto(request):
    empresa_actual = request.tenant

    if request.method == 'POST':
        form = CambioProductoForm(request.POST, empresa=empresa_actual)
        if form.is_valid():
            try:
                with transaction.atomic():

                    cambio = form.save(commit=False)
                    cambio.empresa = empresa_actual
                    cambio.usuario = request.user
                    cambio.save()

                    bodega_principal = Bodega.objects.principal(empresa_actual)

                    # Prevenir duplicados: usar get_or_create para movimientos de cambio
                    mov_entrada, _ = MovimientoInventario.objects.get_or_create(
                        empresa=empresa_actual,
                        producto=cambio.producto_entrante,
                        bodega=bodega_principal,
                        tipo_movimiento='ENTRADA_CAMBIO',
                        documento_referencia=f"Cambio #{cambio.pk}",
                        defaults={
                            'cantidad': cambio.cantidad_entrante,
                            'usuario': request.user
                        }
                    )

                    # Prevenir duplicados: usar get_or_create para movimientos de salida
                    mov_salida, _ = MovimientoInventario.objects.get_or_create(
                        empresa=empresa_actual,
                        producto=cambio.producto_saliente,
                        bodega=bodega_principal,
                        tipo_movimiento='SALIDA_CAMBIO',
                        documento_referencia=f"Cambio #{cambio.pk}",
                        defaults={
                            'cantidad': -abs(cambio.cantidad_saliente),
                            'usuario': request.user
                        }
                    )
                

                pdf_url = reverse('bodega:generar_pdf_cambio_producto', kwargs={'pk': cambio.pk})
                

                mensaje_exito = mark_safe(
                    f"Cambio #{cambio.pk} realizado exitosamente. El stock ha sido actualizado. "
                    f"<a href='{pdf_url}' class='alert-link' target='_blank'><strong>Imprimir Comprobante</strong></a>"
                )
                

                messages.success(request, mensaje_exito)


                return redirect('bodega:realizar_cambio_producto')

            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado al procesar el cambio: {e}")
    else:
        form = CambioProductoForm(empresa=empresa_actual)

    context = {
        'form': form,
        'titulo': 'Realizar Cambio de Producto'
    }
    return render(request, 'bodega/cambio_producto.html', context)

@login_required
@permission_required('bodega.can_manage_product_changes', login_url='core:acceso_denegado')
def historial_cambios_producto(request):
    empresa_actual = request.tenant
    
    # Query base optimizada para evitar consultas repetitivas en la plantilla
    cambios_list = CambioProducto.objects.filter(
        empresa=empresa_actual
    ).select_related(
        'usuario', 'producto_entrante', 'producto_saliente'
    ).order_by('-fecha_hora')

    # Aplicar filtros de búsqueda desde la URL (GET parameters)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    producto_q = request.GET.get('producto_q', '').strip()

    if fecha_inicio:
        cambios_list = cambios_list.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        cambios_list = cambios_list.filter(fecha_hora__date__lte=fecha_fin)
    if producto_q:
        # Busca la referencia en el producto entrante O en el saliente
        cambios_list = cambios_list.filter(
            Q(producto_entrante__referencia__icontains=producto_q) |
            Q(producto_saliente__referencia__icontains=producto_q)
        )

    # Paginación para manejar grandes volúmenes de datos
    paginator = Paginator(cambios_list, 20) # 20 cambios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'cambios_page': page_obj,
        'titulo': 'Historial de Cambios de Producto',
        'filtros_activos': request.GET # Para mantener los valores en el formulario de filtro
    }
    return render(request, 'bodega/historial_cambios.html', context)



@login_required
@permission_required('bodega.can_manage_product_changes', login_url='core:acceso_denegado')
def generar_pdf_cambio_producto(request, pk):
    empresa_actual = request.tenant
    cambio = get_object_or_404(
        CambioProducto.objects.select_related(
            'usuario', 'producto_entrante', 'producto_saliente'
        ),
        pk=pk,
        empresa=empresa_actual
    )

    context_pdf = {
        'cambio': cambio,
        'logo_base64': get_logo_base_64_despacho(empresa=empresa_actual),
        'fecha_generacion': timezone.now(),
    }

    filename = f"Comprobante_Cambio_{empresa_actual.pk}_{cambio.pk}"
    
    return render_pdf_weasyprint(
        request,
        'bodega/cambio_producto_pdf.html',
        context_pdf,
        filename_prefix=filename
    )
    
    
@login_required
@permission_required('bodega.view_movimientoinventario', login_url='core:acceso_denegado')
def buscar_informe_movimiento(request):
    """
    Muestra un formulario para seleccionar un producto y luego redirige
    al nuevo informe de movimientos detallado (Kardex).
    """
    empresa_actual = request.tenant

    if 'producto_id' in request.GET:
        producto_id = request.GET.get('producto_id')
        if producto_id:
            return redirect('bodega:informe_movimiento_producto', pk=producto_id)

    # --- CORRECCIÓN CLAVE ---
    # Buscamos todos los productos para pasarlos a la plantilla
    productos_para_selector = Producto.objects.filter(empresa=empresa_actual, activo=True).order_by('referencia')

    context = {
        'titulo': 'Buscar Kardex de Producto',
        'productos': productos_para_selector, # <-- Se los enviamos a la plantilla
    }
    return render(request, 'bodega/buscar_informe_movimiento.html', context)

@login_required
# Asegúrate que el permiso sea el adecuado, ej: 'bodega.view_movimientoinventario'
@permission_required('bodega.view_movimientoinventario', login_url='core:acceso_denegado')
def informe_movimiento_producto(request, pk):
    """
    Muestra el historial de movimientos (Kardex) para un único producto y
    permite descargar el informe en PDF.
    """
    empresa_actual = request.tenant
    producto = get_object_or_404(
        Producto,
        pk=pk,
        empresa=empresa_actual
    )

    # Desglose de stock actual por bodega (todas las bodegas de la empresa,
    # incluyendo las que tienen 0 para que se note dónde NO hay existencias).
    stock_por_bodega_map = {
        fila['bodega_id']: fila['total']
        for fila in MovimientoInventario.objects.filter(producto=producto).values('bodega_id').annotate(total=Sum('cantidad'))
    }
    stock_por_bodega = [
        {'bodega': b, 'cantidad': stock_por_bodega_map.get(b.pk, 0) or 0}
        for b in Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre')
    ]

    bodega_filtro_id = request.GET.get('bodega_id', '').strip()

    # Obtenemos los movimientos para este producto, ordenados del más reciente al más antiguo
    movimientos_queryset = MovimientoInventario.objects.filter(
        producto=producto
    ).select_related('usuario', 'bodega').order_by('-fecha_hora')

    if bodega_filtro_id:
        movimientos_queryset = movimientos_queryset.filter(bodega_id=bodega_filtro_id)
        # Si se filtra por una bodega específica, el saldo corriente debe
        # partir del stock actual EN ESA BODEGA, no del stock global del producto.
        saldo = stock_por_bodega_map.get(int(bodega_filtro_id), 0) or 0
    else:
        saldo = producto.stock_actual

    # Calcular el saldo corriendo para cada movimiento (Kardex)
    movimientos_con_saldo = []
    for movimiento in movimientos_queryset:
        movimientos_con_saldo.append({
            'movimiento': movimiento,
            'saldo_anterior': saldo,
            'saldo_nuevo': saldo - movimiento.cantidad
        })
        saldo -= movimiento.cantidad

    # Invertimos la lista para mostrar del más antiguo al más reciente en la plantilla
    movimientos_con_saldo.reverse()


    context = {
        'producto': producto,
        'movimientos_list': movimientos_con_saldo,
        'stock_actual_verificacion': producto.stock_actual,
        'stock_por_bodega': stock_por_bodega,
        'bodegas_list': Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre'),
        'bodega_filtro_id': bodega_filtro_id,
        'titulo': f"Kardex de Movimientos - {producto.referencia}",
        'empresa_actual': empresa_actual,
        'fecha_generacion': timezone.now(),
        'logo_base64': get_logo_base_64_despacho(empresa=empresa_actual),
    }

    # Decidir si generar PDF o renderizar HTML
    if request.GET.get('format') == 'pdf':
        filename = f"Kardex_{empresa_actual.pk}_{producto.referencia}_{timezone.now().strftime('%Y%m%d')}"
        return render_pdf_weasyprint(
            request,
            'bodega/informe_movimiento_producto_pdf.html', # Plantilla para el PDF
            context,
            filename_prefix=filename
        )

    # Por defecto, mostrar la vista en pantalla
    return render(request, 'bodega/informe_movimiento_producto.html', context)


@login_required
@permission_required('productos.view_inventory_report', login_url='core:acceso_denegado')
def consulta_stock_bodega(request):
    """
    Consulta rápida: busca un producto por referencia, nombre o código de
    barras y muestra al instante su stock en cada bodega, sin cargar todo
    el historial de movimientos (a diferencia del Kardex).
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        messages.error(request, "No se ha podido identificar la empresa.")
        return redirect('core:index')

    query = request.GET.get('q', '').strip()
    resultados = []

    if query:
        productos = Producto.objects.filter(
            empresa=empresa_actual, activo=True
        ).filter(
            Q(referencia__icontains=query) | Q(nombre__icontains=query) | Q(codigo_barras__icontains=query)
        ).select_related('ubicacion').order_by('referencia', 'color', 'talla')[:50]

        bodegas = list(Bodega.objects.filter(empresa=empresa_actual).order_by('orden', 'nombre'))
        producto_ids = [p.pk for p in productos]

        stock_map = {}
        if producto_ids:
            filas_stock = MovimientoInventario.objects.filter(
                producto_id__in=producto_ids
            ).values('producto_id', 'bodega_id').annotate(total=Sum('cantidad'))
            for fila in filas_stock:
                stock_map[(fila['producto_id'], fila['bodega_id'])] = fila['total'] or 0

        for producto in productos:
            desglose = [
                {'bodega': b, 'cantidad': stock_map.get((producto.pk, b.pk), 0) or 0}
                for b in bodegas
            ]
            resultados.append({
                'producto': producto,
                'desglose': desglose,
                'total': sum(fila['cantidad'] for fila in desglose),
            })

    context = {
        'titulo': 'Consulta Rápida de Stock por Bodega',
        'query': query,
        'resultados': resultados,
    }
    return render(request, 'bodega/consulta_stock_bodega.html', context)


@login_required
@permission_required('bodega.add_movimientoinventario', raise_exception=True)
def vista_ingreso_produccion(request):
    """
    Permite al personal de bodega ingresar stock por lotes de Referencia+Color.
    """
    empresa_actual = getattr(request, 'tenant', None)

    if request.method == 'POST':
        with transaction.atomic(): # Usamos una transacción para asegurar que todo o nada se guarde
            productos_actualizados = 0
            bodega_principal = Bodega.objects.principal(empresa_actual)
            # Buscamos todos los campos que empiezan con 'cantidad_'
            for key, value in request.POST.items():
                if key.startswith('cantidad_'):
                    try:
                        producto_id = int(key.split('_')[1])
                        cantidad = int(value)
                    except (ValueError, TypeError, IndexError):
                        continue

                    if cantidad > 0:
                        producto = get_object_or_404(Producto, pk=producto_id, empresa=empresa_actual)

                        MovimientoInventario.objects.create(
                            empresa=empresa_actual,
                            producto=producto,
                            bodega=bodega_principal,
                            cantidad=cantidad,
                            tipo_movimiento='ENTRADA_PRODUCCION',
                            documento_referencia='Ingreso Producción Lote',
                            usuario=request.user,
                            notas=f'Se ingresaron {cantidad} unidades desde producción.'
                        )
                        
                        producto.permitir_preventa = False
                        producto.save(update_fields=['permitir_preventa'])
                        productos_actualizados += 1
            
            if productos_actualizados > 0:
                messages.success(request, f"Se actualizó el stock para {productos_actualizados} variante(s) de talla.")
            else:
                messages.warning(request, "No se ingresaron cantidades para ninguna talla.")
        
        return redirect('bodega:ingreso_produccion')

    # Para el método GET, ahora pasamos las referencias y colores únicos
    referencias_en_preventa = Producto.objects.filter(
        empresa=empresa_actual,
        permitir_preventa=True
    ).values('referencia', 'color').distinct().order_by('referencia', 'color')

    context = {
        'titulo': 'Ingresar Stock de Producción (por Lote)',
        'referencias_seleccionables': list(referencias_en_preventa),
    }
    return render(request, 'bodega/ingreso_produccion.html', context)

@login_required
def api_get_tallas_para_ingreso(request):
    """
    API que devuelve las tallas de una Referencia+Color que están
    marcadas como 'permitir_preventa'.
    """
    referencia = request.GET.get('ref')
    color = request.GET.get('color')
    empresa_actual = getattr(request, 'tenant', None)

    if not (referencia and color and empresa_actual):
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    # El slug '-' representa los productos sin color asignado
    color_filtro = None if color == '-' else color

    productos = Producto.objects.filter(
        empresa=empresa_actual,
        referencia=referencia,
        color=color_filtro,
        permitir_preventa=True
    ).order_by('talla')

    # Preparamos los datos para enviarlos como JSON
    tallas_data = [
        {
            'pk': p.id,
            'talla': p.talla,
            'stock_actual': p.stock_actual, # Útil para ver la demanda (ej: -5)
        }
        for p in productos
    ]

    return JsonResponse({'tallas': tallas_data})

@login_required
def api_get_tallas_para_ajuste(request):
    """
    API que devuelve TODAS las tallas de una Referencia+Color para
    realizar un ajuste masivo de inventario.
    """
    referencia = request.GET.get('ref')
    color = request.GET.get('color')
    empresa_actual = getattr(request, 'tenant', None)

    if not (referencia and color and empresa_actual):
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    color_filtro = None if color == '-' else color

    # A diferencia de la otra API, aquí buscamos TODAS las tallas, no solo las de preventa
    productos = Producto.objects.filter(
        empresa=empresa_actual,
        referencia=referencia,
        color=color_filtro
    ).order_by('talla')

    tallas_data = [
        {
            'pk': p.id,
            'talla': p.talla,
            'stock_actual': p.stock_actual,
        }
        for p in productos
    ]

    return JsonResponse({'tallas': tallas_data})

@login_required
@permission_required('bodega.add_movimientoinventario', raise_exception=True)
def vista_ajuste_masivo_inventario(request):
    """
    Permite realizar ajustes de inventario masivos para todas las tallas
    de una referencia y color específicos.
    """
    empresa_actual = getattr(request, 'tenant', None)

    if request.method == 'POST':
        tipo_movimiento = request.POST.get('tipo_movimiento')
        documento_referencia = request.POST.get('documento_referencia', 'Ajuste Masivo')

        # Validamos que se haya seleccionado un tipo de movimiento válido
        if not tipo_movimiento or tipo_movimiento not in [choice[0] for choice in MovimientoInventario.TIPO_MOVIMIENTO_CHOICES]:
            messages.error(request, "Debes seleccionar un tipo de movimiento válido.")
            return redirect('bodega:ajuste_masivo_inventario')

        with transaction.atomic():
            productos_ajustados = 0
            bodega_principal = Bodega.objects.principal(empresa_actual)
            for key, value in request.POST.items():
                if key.startswith('cantidad_'):
                    try:
                        producto_id = int(key.split('_')[1])
                        # Permitimos cantidades negativas y positivas
                        cantidad = int(value) if value else 0
                    except (ValueError, TypeError, IndexError):
                        continue

                    if cantidad != 0:
                        producto = get_object_or_404(Producto, pk=producto_id, empresa=empresa_actual)

                        MovimientoInventario.objects.create(
                            empresa=empresa_actual,
                            producto=producto,
                            bodega=bodega_principal,
                            cantidad=cantidad,
                            tipo_movimiento=tipo_movimiento,
                            documento_referencia=documento_referencia,
                            usuario=request.user
                        )
                        productos_ajustados += 1
            
            if productos_ajustados > 0:
                messages.success(request, f"Se aplicó el ajuste de inventario para {productos_ajustados} variante(s) de talla.")
            else:
                messages.warning(request, "No se ingresaron cantidades para ajustar.")
        
        return redirect('bodega:ajuste_masivo_inventario')

    # Para el método GET, pasamos las referencias/colores y los tipos de movimiento
    referencias_unicas = Producto.objects.filter(
        empresa=empresa_actual, activo=True
    ).values('referencia', 'color').distinct().order_by('referencia', 'color')

    context = {
        'titulo': 'Ajuste Masivo de Inventario',
        'referencias_seleccionables': list(referencias_unicas),
        'tipos_movimiento': MovimientoInventario.TIPO_MOVIMIENTO_CHOICES,
    }
    return render(request, 'bodega/ajuste_masivo_inventario.html', context)
    

@login_required
def visibilidad_productos_view(request):
    """Vista web para que el bodeguero busque y vea qué ocultar"""
    # 1. Filtro MULTI-INQUILINO: Solo los productos de la empresa actual
    empresa_actual = getattr(request, 'tenant', None) or request.user.empresa
    qs = Producto.objects.filter(empresa=empresa_actual)
    
    # 2. Buscador
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(referencia__icontains=q) | qs.filter(descripcion__icontains=q)
        
    # 3. Agrupamos para no repetir referencias
    referencias_unicas = {}
    for p in qs:
        if p.referencia not in referencias_unicas:
            referencias_unicas[p.referencia] = p
            
    context = {
        'referencias': list(referencias_unicas.values())
    }
    return render(request, 'bodega/visibilidad_productos.html', context)

@login_required
def toggle_visibilidad_view(request):
    """Endpoint AJAX para encender/apagar todas las tallas de una referencia"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            referencia = data.get('referencia')
            
            # Filtro MULTI-INQUILINO por seguridad
            empresa_actual = getattr(request, 'tenant', None) or request.user.empresa
            productos = Producto.objects.filter(referencia=referencia, empresa=empresa_actual)
            
            if productos.exists():
                producto_base = productos.first()
                nuevo_estado = not producto_base.oculto_para_standar
                
                # Actualizamos todas las tallas de esta referencia de un golpe
                productos.update(oculto_para_standar=nuevo_estado)
                
                return JsonResponse({'status': 'ok', 'oculto': nuevo_estado})
            else:
                return JsonResponse({'status': 'error', 'msg': 'Referencia no encontrada'}, status=404)
                
        except Exception as e:
            return JsonResponse({'status': 'error', 'msg': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'msg': 'Método no permitido'}, status=405)


# ============================================================
# CRUD de Bodegas
# ============================================================

class BodegaListView(TenantAwareMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Bodega
    template_name = 'bodega/bodega_lista.html'
    context_object_name = 'bodegas_list'
    permission_required = 'bodega.view_bodega'
    login_url = 'core:acceso_denegado'

    def get_queryset(self):
        # No usamos el bypass de superuser de TenantAwareMixin: las bodegas son
        # siempre de UNA empresa, incluso para un superusuario administrando
        # una empresa específica. De lo contrario se mezclarían las bodegas
        # (incluida la Principal) de todas las empresas del sistema.
        return Bodega.objects.filter(empresa=self.request.tenant).select_related('responsable').order_by('orden', 'nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Bodegas'
        return context


class BodegaCreateView(TenantAwareMixin, LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Bodega
    form_class = BodegaForm
    template_name = 'bodega/bodega_form.html'
    success_url = reverse_lazy('bodega:lista_bodegas')
    success_message = "Bodega '%(nombre)s' creada exitosamente."
    permission_required = 'bodega.add_bodega'
    login_url = 'core:acceso_denegado'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.request.tenant
        return kwargs

    def form_valid(self, form):
        # TenantAwareMixin.form_valid usa hasattr(form.instance, 'empresa') para
        # decidir si asigna la empresa, pero Bodega.empresa es un FK obligatorio:
        # acceder a él en una instancia sin guardar lanza RelatedObjectDoesNotExist
        # (que Django hace heredar de AttributeError), así que hasattr() da False
        # y la mixin nunca llega a asignarla. Se asigna aquí explícitamente.
        if not form.instance.pk:
            form.instance.empresa = self.request.tenant
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Bodega'
        return context


class BodegaUpdateView(TenantAwareMixin, LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Bodega
    form_class = BodegaForm
    template_name = 'bodega/bodega_form.html'
    success_url = reverse_lazy('bodega:lista_bodegas')
    success_message = "Bodega '%(nombre)s' actualizada exitosamente."
    permission_required = 'bodega.change_bodega'
    login_url = 'core:acceso_denegado'

    def get_queryset(self):
        # Igual que en BodegaListView: nunca saltarse el filtro de empresa,
        # ni para superusuarios, para no permitir editar la bodega de otra
        # empresa cambiando el ID en la URL.
        return Bodega.objects.filter(empresa=self.request.tenant)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.request.tenant
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Bodega: {self.object.nombre}'
        return context


@login_required
@permission_required('bodega.change_bodega', login_url='core:acceso_denegado')
@require_POST
def toggle_bodega_activa(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    bodega = get_object_or_404(Bodega, pk=pk, empresa=empresa_actual)
    if bodega.es_principal:
        messages.error(request, "La Bodega Principal no se puede desactivar.")
    else:
        bodega.activa = not bodega.activa
        bodega.save(update_fields=['activa'])
        estado_str = 'activada' if bodega.activa else 'desactivada'
        messages.success(request, f"Bodega '{bodega.nombre}' {estado_str} exitosamente.")
    return redirect('bodega:lista_bodegas')


def _referencias_bodega(bodega):
    """
    Cuenta los registros que dependen de esta bodega (todos con on_delete=PROTECT,
    salvo los accesos que se borran en cascada). Si algún contador es mayor a 0,
    la bodega no se puede eliminar sin perder historial.
    """
    return {
        'Productos ubicados aquí': bodega.productos.count(),
        'Movimientos de inventario': bodega.movimientos.count(),
        'Ingresos a bodega': bodega.ingresos.count(),
        'Conteos de inventario': bodega.conteos.count(),
        'Salidas internas (origen)': bodega.salidas_internas_origen.count(),
        'Traslados (origen)': bodega.traslados_salientes.count(),
        'Traslados (destino)': bodega.traslados_entrantes.count(),
    }


@login_required
@permission_required('bodega.delete_bodega', login_url='core:acceso_denegado')
def eliminar_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    bodega = get_object_or_404(Bodega, pk=pk, empresa=empresa_actual)

    if bodega.es_principal:
        messages.error(request, "La Bodega Principal no se puede eliminar.")
        return redirect('bodega:lista_bodegas')

    referencias = _referencias_bodega(bodega)
    referencias_con_datos = {k: v for k, v in referencias.items() if v > 0}

    if request.method == 'POST':
        if referencias_con_datos:
            messages.error(
                request,
                f"No se puede eliminar la bodega '{bodega.nombre}' porque todavía tiene historial asociado. "
                f"Desactívala en su lugar si ya no la vas a usar."
            )
            return redirect('bodega:lista_bodegas')

        AccesoBodega.objects.filter(bodega=bodega).delete()
        nombre = bodega.nombre
        bodega.delete()
        messages.success(request, f"Bodega '{nombre}' eliminada exitosamente.")
        return redirect('bodega:lista_bodegas')

    context = {
        'bodega': bodega,
        'referencias_con_datos': referencias_con_datos,
        'titulo': f'Eliminar Bodega: {bodega.nombre}',
    }
    return render(request, 'bodega/bodega_confirm_delete.html', context)


# ============================================================
# Administración de Accesos por Bodega
# ============================================================

@login_required
@permission_required('bodega.view_accesobodega', login_url='core:acceso_denegado')
def lista_accesos_bodega(request):
    empresa_actual = getattr(request, 'tenant', None)
    accesos = AccesoBodega.objects.filter(bodega__empresa=empresa_actual).select_related('usuario', 'bodega').order_by('bodega__nombre', 'usuario__username')
    context = {
        'accesos_list': accesos,
        'titulo': 'Accesos a Bodegas por Usuario',
    }
    return render(request, 'bodega/acceso_bodega_lista.html', context)


@login_required
@permission_required('bodega.add_accesobodega', login_url='core:acceso_denegado')
def crear_acceso_bodega(request):
    empresa_actual = getattr(request, 'tenant', None)
    if request.method == 'POST':
        form = AccesoBodegaForm(request.POST, empresa=empresa_actual)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Acceso a bodega registrado exitosamente.")
                return redirect('bodega:lista_accesos_bodega')
            except IntegrityError:
                messages.error(request, "Ese usuario ya tiene un acceso configurado para esa bodega. Edítalo en vez de crear uno nuevo.")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = AccesoBodegaForm(empresa=empresa_actual)

    context = {
        'form': form,
        'titulo': 'Nuevo Acceso a Bodega',
    }
    return render(request, 'bodega/acceso_bodega_form.html', context)


class AccesoBodegaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = AccesoBodega
    template_name = 'bodega/acceso_bodega_confirm_delete.html'
    success_url = reverse_lazy('bodega:lista_accesos_bodega')
    permission_required = 'bodega.delete_accesobodega'
    login_url = 'core:acceso_denegado'

    def get_queryset(self):
        empresa_actual = getattr(self.request, 'tenant', None)
        return AccesoBodega.objects.filter(bodega__empresa=empresa_actual)

    def form_valid(self, form):
        messages.success(self.request, "Acceso a bodega eliminado exitosamente.")
        return super().form_valid(form)


# ============================================================
# Traslados entre Bodegas
# ============================================================

@login_required
@permission_required('bodega.view_trasladobodega', login_url='core:acceso_denegado')
def lista_traslados_bodega(request):
    empresa_actual = getattr(request, 'tenant', None)
    traslados = TrasladoBodega.objects.filter(empresa=empresa_actual).select_related(
        'bodega_origen', 'bodega_destino'
    ).order_by('-fecha_hora_creacion')
    context = {
        'traslados_list': traslados,
        'titulo': 'Traslados entre Bodegas',
    }
    return render(request, 'bodega/traslado_lista.html', context)


@login_required
@permission_required('bodega.add_trasladobodega', login_url='core:acceso_denegado')
def api_productos_con_stock_bodega(request):
    """
    Devuelve en JSON los productos (activos, de la empresa actual) que tienen
    stock > 0 en la bodega indicada. Se usa para que, al escoger la Bodega de
    Origen de un traslado, el selector de 'Productos a Trasladar' solo
    ofrezca lo que realmente hay disponible ahí.
    """
    empresa_actual = getattr(request, 'tenant', None)
    bodega_id = request.GET.get('bodega_id')

    if not empresa_actual or not bodega_id:
        return JsonResponse({'productos': []})

    bodega = Bodega.objects.filter(pk=bodega_id, empresa=empresa_actual).first()
    if not bodega:
        return JsonResponse({'productos': []})

    filas_stock = MovimientoInventario.objects.filter(
        bodega=bodega, producto__empresa=empresa_actual, producto__activo=True
    ).values('producto_id').annotate(total=Sum('cantidad')).filter(total__gt=0)

    stock_por_producto = {fila['producto_id']: fila['total'] for fila in filas_stock}

    productos = Producto.objects.filter(pk__in=stock_por_producto.keys()).order_by('referencia', 'color', 'talla')

    data = [
        {
            'id': p.pk,
            'label': f"{p.referencia} - {p.nombre} ({p.color or '-'} - Talla {p.talla or '-'}) | Stock: {stock_por_producto[p.pk]}",
            'stock': stock_por_producto[p.pk],
        }
        for p in productos
    ]
    return JsonResponse({'productos': data})


@login_required
@permission_required('bodega.add_trasladobodega', login_url='core:acceso_denegado')
def crear_traslado_bodega(request):
    """
    El detalle del traslado ya no se arma línea por línea con un formset:
    la plantilla carga (vía AJAX) todos los productos con stock en la
    Bodega de Origen elegida y el usuario marca un check + cantidad por
    cada uno que quiera incluir, para poder trasladar muchas referencias
    de un tirón.
    """
    empresa_actual = getattr(request, 'tenant', None)

    if request.method == 'POST':
        form = TrasladoBodegaForm(request.POST, empresa=empresa_actual)

        # Recolectar las filas marcadas: 'incluir_<producto_id>' presente y
        # 'cantidad_<producto_id>' > 0.
        productos_incluidos = {}
        for key in request.POST.keys():
            if key.startswith('incluir_'):
                try:
                    producto_id = int(key.split('_', 1)[1])
                    cantidad = int(request.POST.get(f'cantidad_{producto_id}', '0') or 0)
                except (ValueError, TypeError, IndexError):
                    continue
                if cantidad > 0:
                    productos_incluidos[producto_id] = cantidad

        if form.is_valid():
            if not productos_incluidos:
                messages.warning(request, "Debes marcar al menos un producto con cantidad para el traslado.")
            else:
                bodega_origen = form.cleaned_data['bodega_origen']
                productos_map = {
                    p.pk: p for p in Producto.objects.filter(empresa=empresa_actual, pk__in=productos_incluidos.keys())
                }

                # Defensa en profundidad: aunque el selector en pantalla ya
                # solo muestra productos con stock en la bodega de origen, se
                # vuelve a validar aquí por si el stock cambió entre que se
                # cargó la página y se envió el formulario.
                errores_stock = []
                for producto_id, cantidad in productos_incluidos.items():
                    producto = productos_map.get(producto_id)
                    if not producto:
                        continue
                    stock_en_origen = MovimientoInventario.objects.filter(
                        producto=producto, bodega=bodega_origen
                    ).aggregate(total=Sum('cantidad'))['total'] or 0
                    if stock_en_origen < cantidad:
                        errores_stock.append(
                            f"'{producto}': disponible {stock_en_origen} en '{bodega_origen.nombre}', solicitado {cantidad}."
                        )

                if errores_stock:
                    for error in errores_stock:
                        messages.error(request, error)
                else:
                    with transaction.atomic():
                        traslado = form.save(commit=False)
                        traslado.empresa = empresa_actual
                        traslado.usuario_creacion = request.user
                        traslado.save()

                        for producto_id, cantidad in productos_incluidos.items():
                            DetalleTrasladoBodega.objects.create(
                                traslado=traslado,
                                producto=productos_map[producto_id],
                                cantidad=cantidad,
                            )

                    messages.success(request, f"Traslado #{traslado.pk} creado como borrador con {len(productos_incluidos)} producto(s). Envíalo cuando esté listo.")
                    return redirect('bodega:detalle_traslado', pk=traslado.pk)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = TrasladoBodegaForm(empresa=empresa_actual)

    context = {
        'form': form,
        'titulo': 'Registrar Traslado entre Bodegas',
    }
    return render(request, 'bodega/traslado_form.html', context)


@login_required
@permission_required('bodega.view_trasladobodega', login_url='core:acceso_denegado')
def detalle_traslado_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    traslado = get_object_or_404(
        TrasladoBodega.objects.select_related('bodega_origen', 'bodega_destino').prefetch_related('detalles__producto'),
        pk=pk, empresa=empresa_actual
    )
    context = {
        'traslado': traslado,
        'titulo': f'Traslado #{traslado.pk}',
    }
    return render(request, 'bodega/traslado_detalle.html', context)


@login_required
@permission_required('bodega.change_trasladobodega', login_url='core:acceso_denegado')
@require_POST
def enviar_traslado_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    traslado = get_object_or_404(
        TrasladoBodega.objects.select_related('bodega_origen', 'bodega_destino').prefetch_related('detalles__producto'),
        pk=pk, empresa=empresa_actual
    )

    if traslado.estado != 'BORRADOR':
        messages.error(request, f"El traslado #{traslado.pk} ya fue enviado o no está en estado Borrador.")
        return redirect('bodega:detalle_traslado', pk=traslado.pk)

    # Validar stock disponible en la bodega de origen para cada producto
    for detalle in traslado.detalles.all():
        stock_en_origen = MovimientoInventario.objects.filter(
            producto=detalle.producto, bodega=traslado.bodega_origen
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        if stock_en_origen < detalle.cantidad:
            messages.error(
                request,
                f"Stock insuficiente en '{traslado.bodega_origen.nombre}' para '{detalle.producto}'. "
                f"Disponible: {stock_en_origen}, requerido: {detalle.cantidad}."
            )
            return redirect('bodega:detalle_traslado', pk=traslado.pk)

    with transaction.atomic():
        for detalle in traslado.detalles.all():
            MovimientoInventario.objects.get_or_create(
                empresa=empresa_actual,
                producto=detalle.producto,
                bodega=traslado.bodega_origen,
                tipo_movimiento='SALIDA_TRASLADO',
                documento_referencia=f"Traslado #{traslado.pk} - {detalle.producto.pk}",
                defaults={
                    'cantidad': -detalle.cantidad,
                    'usuario': request.user,
                    'notas': f"Salida por Traslado #{traslado.pk} hacia {traslado.bodega_destino.nombre}",
                }
            )

        traslado.usuario_envio = request.user
        traslado.fecha_hora_envio = timezone.now()

        if traslado.bodega_destino.requiere_confirmacion_recepcion:
            traslado.estado = 'EN_TRANSITO'
            traslado.save()
            messages.success(request, f"Traslado #{traslado.pk} enviado. Queda 'En Tránsito' hasta que se confirme la recepción en '{traslado.bodega_destino.nombre}'.")
        else:
            for detalle in traslado.detalles.all():
                MovimientoInventario.objects.get_or_create(
                    empresa=empresa_actual,
                    producto=detalle.producto,
                    bodega=traslado.bodega_destino,
                    tipo_movimiento='ENTRADA_TRASLADO',
                    documento_referencia=f"Traslado #{traslado.pk} - {detalle.producto.pk}",
                    defaults={
                        'cantidad': detalle.cantidad,
                        'usuario': request.user,
                        'notas': f"Entrada por Traslado #{traslado.pk} desde {traslado.bodega_origen.nombre}",
                    }
                )
            traslado.estado = 'RECIBIDO'
            traslado.usuario_recepcion = request.user
            traslado.fecha_hora_recepcion = timezone.now()
            traslado.save()
            messages.success(request, f"Traslado #{traslado.pk} enviado y recibido automáticamente en '{traslado.bodega_destino.nombre}'.")

    return redirect('bodega:detalle_traslado', pk=traslado.pk)


@login_required
@permission_required('bodega.change_trasladobodega', login_url='core:acceso_denegado')
@require_POST
def confirmar_recepcion_traslado_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    traslado = get_object_or_404(
        TrasladoBodega.objects.select_related('bodega_origen', 'bodega_destino').prefetch_related('detalles__producto'),
        pk=pk, empresa=empresa_actual
    )

    if traslado.estado != 'EN_TRANSITO':
        messages.error(request, f"El traslado #{traslado.pk} no está pendiente de recepción.")
        return redirect('bodega:detalle_traslado', pk=traslado.pk)

    with transaction.atomic():
        for detalle in traslado.detalles.all():
            MovimientoInventario.objects.get_or_create(
                empresa=empresa_actual,
                producto=detalle.producto,
                bodega=traslado.bodega_destino,
                tipo_movimiento='ENTRADA_TRASLADO',
                documento_referencia=f"Traslado #{traslado.pk} - {detalle.producto.pk}",
                defaults={
                    'cantidad': detalle.cantidad,
                    'usuario': request.user,
                    'notas': f"Entrada por Traslado #{traslado.pk} desde {traslado.bodega_origen.nombre}",
                }
            )
        traslado.estado = 'RECIBIDO'
        traslado.usuario_recepcion = request.user
        traslado.fecha_hora_recepcion = timezone.now()
        traslado.save()

    messages.success(request, f"Recepción del Traslado #{traslado.pk} confirmada en '{traslado.bodega_destino.nombre}'.")
    return redirect('bodega:detalle_traslado', pk=traslado.pk)


@login_required
@permission_required('bodega.change_trasladobodega', login_url='core:acceso_denegado')
@require_POST
def anular_traslado_bodega(request, pk):
    empresa_actual = getattr(request, 'tenant', None)
    traslado = get_object_or_404(TrasladoBodega, pk=pk, empresa=empresa_actual)

    if traslado.estado != 'BORRADOR':
        messages.error(request, f"Solo se puede anular un traslado en estado Borrador. El traslado #{traslado.pk} está '{traslado.get_estado_display()}'.")
        return redirect('bodega:detalle_traslado', pk=traslado.pk)

    traslado.estado = 'ANULADO'
    traslado.save(update_fields=['estado'])
    messages.success(request, f"Traslado #{traslado.pk} anulado.")
    return redirect('bodega:detalle_traslado', pk=traslado.pk)