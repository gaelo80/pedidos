from io import BytesIO
from django.contrib.auth.decorators import login_required, user_passes_test
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.authentication import SessionAuthentication
from rest_framework.permissions import IsAuthenticated
from productos.models import Producto, Color, normalizar_nombre_color
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter


NO_COLOR_SLUG = '-'

@api_view(['GET'])
@permission_classes([IsAuthenticated]) # Requiere autenticación
def get_colores_por_referencia(request, ref):
    """
    Devuelve una lista de colores únicos ({valor, display}) para una referencia dada,
    manejando el caso 'Sin Color' (None o vacío en BD) con el slug NO_COLOR_SLUG.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Response({"error": "No se pudo identificar la empresa."}, status=403)    

    colores_qs = Producto.objects.filter(
        empresa=empresa_actual,
        referencia=ref,
        activo=True
    ).values_list('color__nombre', flat=True).distinct().order_by('color__nombre')

    colores_procesados = set() # Para evitar duplicados si hay '' y None tratados igual
    respuesta = []
    has_no_color = False

    for color in colores_qs:
        if color is None or color == '':
            if not has_no_color: 
                has_no_color = True
        elif color not in colores_procesados:
            respuesta.append({'valor': color, 'display': color})
            colores_procesados.add(color)
  
    if has_no_color:
        respuesta.append({'valor': NO_COLOR_SLUG, 'display': 'Sin Color'})
    
    respuesta.sort(key=lambda x: x['display'])
    return Response(respuesta)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_tallas_por_ref_color(request, ref, color_slug):
    """
    Devuelve tallas, IDs, precio y stock para una ref+color,
    filtrando por la empresa del usuario autenticado.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Response({"error": "No se pudo identificar la empresa."}, status=403)
    
    # Determinar el filtro de color en la BD
    filtro_color = {'color__isnull': True} if color_slug == NO_COLOR_SLUG else {'color__nombre': color_slug}

    # Buscar variantes activas que coincidan y tengan talla
    variantes = Producto.objects.filter(
        empresa=empresa_actual,
        referencia=ref,
        activo=True,
        talla__isnull=False,
        **filtro_color
    )

    # Convertir a lista para poder ordenar y acceder al precio
    variantes_lista = list(variantes)

    if not variantes_lista:
        return Response({"precio_venta": 0, "variantes": []})

    def sort_key(producto):
        try: 
            return float(producto.talla)
        except (ValueError, TypeError): 
            return float('inf')

    variantes_ordenadas = sorted(variantes_lista, key=sort_key)

    precio_grupo = variantes_ordenadas[0].precio_venta

    # Un vendedor estándar solo debe ver el stock que está en bodegas
    # habilitadas para su canal (o que tenga como excepción individual);
    # Admin/Bodega/Online siguen viendo el stock real total.
    usuario = request.user
    es_admin_o_especial = (
        usuario.is_superuser or
        usuario.groups.filter(name__icontains='bodega').exists() or
        usuario.groups.filter(name__icontains='online').exists()
    )

    def _stock_para_frontend(v):
        if es_admin_o_especial:
            return v.stock_actual
        return v.stock_disponible_para_canal('ESTANDAR', usuario=usuario)

    respuesta = {
        'precio_venta': precio_grupo,
        'variantes': [
            {
                'id': v.id,
                'talla': v.talla,
                'stock_actual': _stock_para_frontend(v),
                # --- LÍNEA AÑADIDA ---
                # Enviamos el valor del nuevo campo al frontend.
                'permitir_pedido_sin_stock': v.permitir_preventa
            } for v in variantes_ordenadas
        ]
    }
    return Response(respuesta)


@api_view(['GET'])
@permission_classes([IsAuthenticated]) 
def buscar_productos_api(request):
    """
    Busca productos por término para Select2, filtrando por empresa
    y reglas de visibilidad para vendedores estándar.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return JsonResponse({'results': []}) 
    
    term = request.GET.get('term', '').strip()
    if len(term) < 2:
        return JsonResponse({'results': []})
            
    # 1. Consulta base
    queryset = Producto.objects.filter(
        empresa=empresa_actual,
        activo=True
    )

    # --- 2. CANDADO DE VISIBILIDAD MULTI-ROL ---
    usuario = request.user
    
    # Comprobamos si el usuario es Admin o si su grupo contiene la palabra 'online' o 'bodega'
    es_admin_o_especial = (
        usuario.is_superuser or 
        usuario.groups.filter(name__icontains='bodega').exists() or 
        usuario.groups.filter(name__icontains='online').exists()
    )
    
    # Si es Vendedor Estándar (NO admin, NO bodega, NO online), bloqueamos los ocultos
    if not es_admin_o_especial:
        queryset = queryset.filter(oculto_para_standar=False)
    # ------------------------------------------

    # 3. Filtrar por término
    queryset = queryset.filter(
        Q(referencia__icontains=term) | Q(nombre__icontains=term) | Q(codigo_barras__icontains=term)
    )[:20]

    results = []
    for prod in queryset:
        texto_opcion = f"{prod.referencia} - {prod.nombre or ''}"
        if prod.color: texto_opcion += f" / {prod.color}"
        if prod.talla: texto_opcion += f" / Talla: {prod.talla}"
        results.append({'id': prod.pk, 'text': texto_opcion})
    
    return JsonResponse({'results': results})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def buscar_referencias_api(request):
    """
    Busca referencias únicas para Select2, filtrando por la empresa
    y aplicando reglas de visibilidad.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return JsonResponse({'results': []})

    term = request.GET.get('term', '').strip()
    if len(term) < 1:
        return JsonResponse({'results': []})

    # 1. Consulta base
    referencias_qs = Producto.objects.filter(
        empresa=empresa_actual,
        referencia__icontains=term,
        activo=True
    )

    # --- 2. CANDADO DE VISIBILIDAD MULTI-ROL ---
    usuario = request.user
    
    es_admin_o_especial = (
        usuario.is_superuser or 
        usuario.groups.filter(name__icontains='bodega').exists() or 
        usuario.groups.filter(name__icontains='online').exists()
    )
    
    if not es_admin_o_especial:
        referencias_qs = referencias_qs.filter(oculto_para_standar=False)
    # ------------------------------------------

    # 3. Agrupar y preparar respuesta
    referencias_qs = referencias_qs.values('referencia').distinct().order_by('referencia')[:20]
    results = [{'id': item['referencia'], 'text': item['referencia']} for item in referencias_qs]

    return JsonResponse({'results': results})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_crear_color(request):
    """
    Crea un color nuevo para la empresa, o devuelve el ya existente si el
    nombre (normalizado) ya está registrado. Se usa desde el formulario de
    creación/edición de productos cuando el color deseado no está en el catálogo.
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Response({"error": "No se pudo identificar la empresa."}, status=403)

    nombre_normalizado = normalizar_nombre_color((request.data.get('nombre') or '').strip())
    if not nombre_normalizado:
        return Response({"error": "El nombre del color es obligatorio."}, status=400)

    color, creado = Color.objects.get_or_create(empresa=empresa_actual, nombre=nombre_normalizado)
    return Response({'id': color.pk, 'nombre': color.nombre, 'creado': creado})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_tallas_por_genero(request):
    """
    Devuelve las tallas que realmente existen hoy en productos de esta
    empresa para un género dado, para ofrecerlas como checkboxes al crear
    un producto nuevo. Se basa en datos reales de Producto.talla, no en la
    configuración de columnas de los PDFs (Empresa.categorias_tallas), que
    puede no coincidir con las tallas reales usadas (ej. UNISEX).
    """
    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Response({"error": "No se pudo identificar la empresa."}, status=403)

    genero = request.GET.get('genero')
    if not genero:
        return Response({"error": "Falta el parámetro 'genero'."}, status=400)

    tallas = Producto.objects.filter(
        empresa=empresa_actual, genero=genero, talla__isnull=False
    ).values_list('talla', flat=True).distinct().order_by('talla')

    return Response({'tallas': list(tallas)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_exportar_catalogo_etiquetas(request):
    """
    Exporta a Excel el catálogo de productos (referencia, nombre, color,
    talla, código de barras) para alimentar un programa externo de
    impresión de etiquetas. Se puede pedir tanto desde el botón "Descargar
    catálogo" del listado de productos (sesión normal del navegador) como
    desde un script externo autenticado con JWT (mismo mecanismo que
    AlmacenDesktop/BodegaDesktop).
    """
    if not request.user.has_perm('productos.view_producto'):
        return Response({"error": "No tienes permiso para ver el catálogo de productos."}, status=403)

    empresa_actual = getattr(request, 'tenant', None)
    if not empresa_actual:
        return Response({"error": "No se pudo identificar la empresa."}, status=403)

    productos = Producto.objects.filter(
        empresa=empresa_actual, activo=True
    ).exclude(codigo_barras__isnull=True).exclude(codigo_barras='').select_related('color').order_by(
        'referencia', 'color__nombre', 'talla'
    )

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = 'Catalogo Etiquetas'

    encabezados = ['Referencia', 'Nombre', 'Color', 'Talla', 'Codigo de Barras']
    hoja.append(encabezados)
    for indice, encabezado in enumerate(encabezados, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = max(len(encabezado) + 2, 14)

    for producto in productos:
        hoja.append([
            producto.referencia,
            producto.nombre,
            producto.color.nombre if producto.color_id else '',
            producto.talla if producto.talla is not None else '',
            producto.codigo_barras,
        ])

    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="catalogo_etiquetas.xlsx"'
    return response